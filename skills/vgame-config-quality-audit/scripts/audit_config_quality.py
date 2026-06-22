#!/usr/bin/env python3
"""Read-only Vgame config quality audit.

This script inspects source Excel files under Config/GameConfig/Datas and
reports structural issues. It never writes to source workbooks.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


KNOWN_BASELINE_DROP_IDS = {
    60004,
    60005,
    1040101,
    1040102,
    1040103,
    10000001,
    10000002,
    10000004,
    10000005,
    50000006,
}


@dataclass
class Issue:
    severity: str
    check: str
    file: str
    row: int | None
    field: str
    value: str
    message: str


class Audit:
    def __init__(self, datas_root: Path, max_issues: int = 1000) -> None:
        self.datas_root = datas_root
        self.max_issues = max_issues
        self.issues: list[Issue] = []
        self.stats: dict[str, int] = {}
        self.path_index: dict[str, list[Path]] = {}
        self.table_inputs: dict[str, list[str]] = {}
        self.item_ids: set[int] = set()
        self.chest_ids: set[int] = set()
        self.drop_ids: set[int] = set()

    def add(
        self,
        severity: str,
        check: str,
        file: str,
        row: int | None,
        field: str,
        value: Any,
        message: str,
    ) -> None:
        if len(self.issues) >= self.max_issues:
            return
        self.issues.append(
            Issue(
                severity=severity,
                check=check,
                file=file,
                row=row,
                field=field,
                value="" if value is None else str(value),
                message=message,
            )
        )

    def rel(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.datas_root))
        except ValueError:
            return str(path)

    def build_path_index(self) -> None:
        self.path_index.clear()
        for path in self.datas_root.rglob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            self.path_index.setdefault(path.name.lower(), []).append(path)
        self.stats["xlsx_files"] = sum(len(v) for v in self.path_index.values())

    def resolve_input(self, filename: str) -> Path | None:
        matches = self.path_index.get(filename.lower(), [])
        if not matches:
            return None
        return matches[0]

    def workbook(self, path: Path):
        return load_workbook(path, read_only=True, data_only=True)

    def first_sheet(self, path: Path):
        wb = self.workbook(path)
        return wb.worksheets[0]

    def headers(self, ws, row: int = 1) -> dict[str, int]:
        out: dict[str, int] = {}
        for idx, cell in enumerate(ws[row], start=1):
            if cell.value is not None:
                out[str(cell.value).strip()] = idx
        return out

    def iter_records(self, path: Path, data_start: int = 5):
        ws = self.first_sheet(path)
        hdr = self.headers(ws, 1)
        for rownum, row in enumerate(ws.iter_rows(min_row=data_start, values_only=False), start=data_start):
            if not any(cell.value is not None for cell in row):
                continue
            record = {name: row[col - 1].value for name, col in hdr.items()}
            yield rownum, record

    def to_int(self, value: Any) -> int | None:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return None
        text = str(value).strip()
        if not text:
            return None
        try:
            return int(text)
        except ValueError:
            try:
                f = float(text)
            except ValueError:
                return None
            return int(f) if f.is_integer() else None

    def split_inputs(self, value: Any) -> list[str]:
        if value is None:
            return []
        return [part.strip() for part in str(value).split(",") if part.strip()]

    def parse_prop_list(self, value: Any) -> tuple[list[tuple[int, int]], list[str]]:
        if value is None or value == "":
            return [], []
        text = str(value).strip()
        if not text or text in {"0", "-1"}:
            return [], []
        props: list[tuple[int, int]] = []
        errors: list[str] = []
        for part in text.split(";"):
            part = part.strip()
            if not part:
                continue
            bits = [b.strip() for b in part.split(",")]
            if len(bits) != 2:
                errors.append(f"bad PropType segment: {part}")
                continue
            item_id = self.to_int(bits[0])
            qty = self.to_int(bits[1])
            if item_id is None or qty is None:
                errors.append(f"bad PropType numeric value: {part}")
                continue
            props.append((item_id, qty))
        return props, errors

    def load_table_registry(self) -> None:
        path = self.datas_root / "__tables__.xlsx"
        if not path.exists():
            self.add("ERROR", "registry", self.rel(path), None, "__tables__", path, "missing __tables__.xlsx")
            return
        ws = self.first_sheet(path)
        hdr = self.headers(ws, 1)
        full_col = hdr.get("full_name")
        input_col = hdr.get("input")
        if not full_col or not input_col:
            self.add("ERROR", "registry", self.rel(path), 1, "header", "", "missing full_name/input headers")
            return
        for row in ws.iter_rows(min_row=4, values_only=False):
            full_name = row[full_col - 1].value
            input_value = row[input_col - 1].value
            if not full_name or not input_value:
                continue
            inputs = self.split_inputs(input_value)
            self.table_inputs[str(full_name).strip()] = inputs
            for filename in inputs:
                matches = self.path_index.get(filename.lower(), [])
                if not matches:
                    self.add("ERROR", "registry", self.rel(path), row[0].row, "input", filename, "registered input file not found under Datas")
                elif len(matches) > 1:
                    self.add("WARN", "registry", self.rel(path), row[0].row, "input", filename, "input filename has multiple matches under Datas")
        self.stats["registered_tables"] = len(self.table_inputs)

    def load_item_ids(self) -> None:
        path = self.resolve_input("Items.xlsx")
        if not path:
            self.add("ERROR", "items", "Items.xlsx", None, "file", "", "Items.xlsx not found")
            return
        for _, rec in self.iter_records(path):
            item_id = self.to_int(rec.get("Id"))
            if item_id is not None:
                self.item_ids.add(item_id)
        self.stats["item_ids"] = len(self.item_ids)

        chest = self.resolve_input("ItemChest.xlsx")
        if chest:
            for _, rec in self.iter_records(chest):
                chest_id = self.to_int(rec.get("ChestId"))
                if chest_id is not None:
                    self.chest_ids.add(chest_id)
        self.stats["chest_ids"] = len(self.chest_ids)

    def load_drop_ids(self) -> None:
        inputs = self.table_inputs.get("Drop", [])
        if not inputs:
            self.add("ERROR", "drop", "__tables__.xlsx", None, "Drop", "", "logical Drop table is not registered")
            return
        for filename in inputs:
            path = self.resolve_input(filename)
            if not path:
                continue
            for _, rec in self.iter_records(path):
                drop_id = self.to_int(rec.get("DropId"))
                if drop_id is not None:
                    self.drop_ids.add(drop_id)
        self.stats["drop_ids"] = len(self.drop_ids)

    def check_drop_sources(self) -> None:
        for filename in self.table_inputs.get("Drop", []):
            path = self.resolve_input(filename)
            if not path:
                continue
            for rownum, rec in self.iter_records(path):
                dmin = self.to_int(rec.get("DropMinNum"))
                dmax = self.to_int(rec.get("DropMaxNum"))
                if dmin is not None and dmax is not None and dmin > dmax:
                    self.add("ERROR", "drop-range", self.rel(path), rownum, "DropMinNum/DropMaxNum", f"{dmin}>{dmax}", "DropMinNum is greater than DropMaxNum")
                imin = self.to_int(rec.get("DropItemNumMin"))
                imax = self.to_int(rec.get("DropItemNumMax"))
                if imin is not None and imax is not None and imin > imax:
                    self.add("ERROR", "drop-range", self.rel(path), rownum, "DropItemNumMin/DropItemNumMax", f"{imin}>{imax}", "DropItemNumMin is greater than DropItemNumMax")
                drop2 = self.to_int(rec.get("DropId2"))
                if drop2 and drop2 not in self.drop_ids:
                    self.add("ERROR", "drop-ref", self.rel(path), rownum, "DropId2", drop2, "DropId2 does not exist in logical Drop")
                item_id = self.to_int(rec.get("DropItem_id"))
                if item_id and item_id not in self.item_ids and item_id not in self.chest_ids:
                    self.add("WARN", "drop-item", self.rel(path), rownum, "DropItem_id", item_id, "DropItem_id not found in Items or ItemChest.ChestId; may be a non-Items reward object")

    def check_drop_ref(self, path: Path, row: int, field: str, value: Any, check: str) -> None:
        drop_id = self.to_int(value)
        if not drop_id:
            return
        if drop_id in self.drop_ids:
            return
        severity = "WARN" if drop_id in KNOWN_BASELINE_DROP_IDS else "ERROR"
        msg = "known baseline missing DropId" if severity == "WARN" else "DropId not found in logical Drop"
        self.add(severity, check, self.rel(path), row, field, drop_id, msg)

    def check_prop_field(self, path: Path, row: int, field: str, value: Any, check: str) -> None:
        props, errors = self.parse_prop_list(value)
        for err in errors:
            self.add("ERROR", check, self.rel(path), row, field, value, err)
        for item_id, _qty in props:
            if item_id not in self.item_ids:
                self.add("ERROR", check, self.rel(path), row, field, value, f"PropType item id {item_id} not found in Items")

    def check_uilevel(self) -> None:
        path = self.resolve_input("UIlevel.xlsx") or self.resolve_input("UILevel.xlsx")
        if not path:
            self.add("WARN", "uilevel", "UIlevel.xlsx", None, "file", "", "UIlevel.xlsx not found")
            return
        fields = ["FirstRewardDrop", "FallRewardDrop", "ThreeStarRewardDrop", "HighRewardDrop"]
        for rownum, rec in self.iter_records(path):
            for field in fields:
                self.check_drop_ref(path, rownum, field, rec.get(field), "uilevel-drop")
            for field in ["Reward", "FirstReward"]:
                self.check_prop_field(path, rownum, field, rec.get(field), "uilevel-preview")

    def check_drawcard(self) -> None:
        path = self.resolve_input("DrawCard.xlsx")
        if not path:
            return
        for rownum, rec in self.iter_records(path):
            for field in ["DropId", "TenDropId"]:
                self.check_drop_ref(path, rownum, field, rec.get(field), "drawcard-drop")
            for field in ["OnePrice", "TenPrice"]:
                self.check_prop_field(path, rownum, field, rec.get(field), "drawcard-price")

    def check_shop_goods(self) -> None:
        path = self.resolve_input("ShopGoods.xlsx")
        if not path:
            return
        for rownum, rec in self.iter_records(path):
            item_id = self.to_int(rec.get("ItemId"))
            if item_id and item_id not in self.item_ids:
                self.add("WARN", "shop-item", self.rel(path), rownum, "ItemId", item_id, "ShopGoods.ItemId not found in Items; may be a non-Items good")
            self.check_prop_field(path, rownum, "Price", rec.get("Price"), "shop-price")

    def check_reward_ref_tables(self) -> None:
        specs = [
            ("Task.xlsx", ["DropId"]),
            ("DailyActiveReward.xlsx", ["RewardId"]),
            ("WeeklyReward.xlsx", ["RewardId"]),
            ("BattlePassReward.xlsx", ["CommonReward", "PassAdditionReward"]),
            ("ActiveReward.xlsx", ["RewardId"]),
            ("PresendMail.xlsx", ["Drop"]),
            ("ChapterStarReward.xlsx", ["DropId"]),
            ("ChapterTasks.xlsx", ["DropId"]),
            ("TowerFirstReward.xlsx", ["DropId"]),
            ("TowerReward.xlsx", ["SeasonDropId"]),
            ("TowerRankReward.xlsx", ["DropId"]),
            ("WeekMaxLayerReward.xlsx", ["CommonDropId"]),
            ("WeekScoreReward.xlsx", ["DropId"]),
            ("SeasonScoreReward.xlsx", ["DropId"]),
            ("SeasonRankReward.xlsx", ["DropId"]),
        ]
        for filename, fields in specs:
            path = self.resolve_input(filename)
            if not path:
                continue
            for rownum, rec in self.iter_records(path):
                for field in fields:
                    self.check_drop_ref(path, rownum, field, rec.get(field), "reward-ref")

    def run(self) -> None:
        self.build_path_index()
        self.load_table_registry()
        self.load_item_ids()
        self.load_drop_ids()
        self.check_drop_sources()
        self.check_uilevel()
        self.check_drawcard()
        self.check_shop_goods()
        self.check_reward_ref_tables()

    def summary(self) -> dict[str, Any]:
        counts: dict[str, int] = {"ERROR": 0, "WARN": 0, "INFO": 0}
        for issue in self.issues:
            counts[issue.severity] = counts.get(issue.severity, 0) + 1
        return {"stats": self.stats, "counts": counts, "issues": [asdict(i) for i in self.issues]}

    def markdown(self) -> str:
        data = self.summary()
        lines = [
            "# Vgame Config Quality Audit",
            "",
            "## Summary",
            "",
            f"- Datas root: `{self.datas_root}`",
            f"- XLSX files indexed: {self.stats.get('xlsx_files', 0)}",
            f"- Registered logical tables: {self.stats.get('registered_tables', 0)}",
            f"- Item IDs: {self.stats.get('item_ids', 0)}",
            f"- Chest IDs: {self.stats.get('chest_ids', 0)}",
            f"- Logical Drop IDs: {self.stats.get('drop_ids', 0)}",
            f"- Errors: {data['counts'].get('ERROR', 0)}",
            f"- Warnings: {data['counts'].get('WARN', 0)}",
            "",
            "## Issues",
            "",
        ]
        if not self.issues:
            lines.append("No issues found by the automated checks.")
            return "\n".join(lines) + "\n"
        lines.append("| Severity | Check | File | Row | Field | Value | Message |")
        lines.append("|---|---|---|---:|---|---|---|")
        for issue in self.issues:
            value = issue.value.replace("|", "\\|")
            msg = issue.message.replace("|", "\\|")
            lines.append(
                f"| {issue.severity} | {issue.check} | `{issue.file}` | {issue.row or ''} | `{issue.field}` | `{value}` | {msg} |"
            )
        return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read-only Vgame config quality audit.")
    parser.add_argument("--datas-root", default=r"D:\Vgame\Config\GameConfig\Datas", help="Path to Config/GameConfig/Datas")
    parser.add_argument("--output", help="Optional Markdown report output path")
    parser.add_argument("--json", action="store_true", help="Print JSON instead of text summary")
    parser.add_argument("--max-issues", type=int, default=1000, help="Maximum issues to collect")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    audit = Audit(Path(args.datas_root), max_issues=args.max_issues)
    audit.run()
    data = audit.summary()
    if args.output:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(audit.markdown(), encoding="utf-8")
    if args.json:
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        print("Vgame config quality audit")
        print(f"Datas root: {audit.datas_root}")
        print(f"Errors: {data['counts'].get('ERROR', 0)}")
        print(f"Warnings: {data['counts'].get('WARN', 0)}")
        if args.output:
            print(f"Report: {args.output}")
    return 1 if data["counts"].get("ERROR", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
