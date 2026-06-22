#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate a read-only full-max character quantification workbook for Vgame."""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_CONFIG_ROOT = Path(r"D:\Vgame\Config\GameConfig\Datas")
DEFAULT_OUTPUT_DIR = Path(r"D:\Vgame\codex_output")
HEADER_FONT_NAME = "黑体"
BODY_FONT_NAME = "微软雅黑"
PEER_DELTA_COLOR_RULES = [
    ("偏强", 0.22, None, "F4CCCC", "9C0006"),
    ("偏高观察", 0.10, 0.22, "FCE4D6", "C65911"),
    ("接近均值", -0.10, 0.10, "E2F0D9", "375623"),
    ("偏弱", None, -0.18, "CFE2F3", "0B5394"),
    ("偏低观察", -0.18, -0.10, "D9EAF7", "1F4E79"),
]


CHARACTER_BENCHMARKS = {
    ("DPS", 4): {
        "HeroId": 1005,
        "Name": "陈单单",
        "Role": "输出SSR标杆",
        "Confidence": "高",
        "Source": "用户指定高置信度角色",
    },
    ("DPS", 3): {
        "HeroId": 1014,
        "Name": "克莉丝汀",
        "Alias": "克里斯汀",
        "Role": "输出SR标杆",
        "Confidence": "高",
        "Source": "用户指定高置信度角色；用户称克里斯汀，配置名为克莉丝汀",
    },
    ("Tank", 4): {
        "HeroId": 1003,
        "Name": "塔拉夏",
        "Role": "坦克SSR标杆",
        "Confidence": "高",
        "Source": "用户指定高置信度角色",
    },
    ("Tank", 3): {
        "HeroId": 1002,
        "Name": "大暑",
        "Role": "坦克SR标杆",
        "Confidence": "高",
        "Source": "用户指定高置信度角色",
    },
    ("Support", 3): {
        "HeroId": 1011,
        "Name": "陈双双",
        "Role": "奶妈SR标杆",
        "Confidence": "高",
        "Source": "用户指定高置信度角色",
    },
}


FIELD_LABELS = {
    "RankInClass": "职业内排名",
    "HeroId": "角色编号",
    "Name": "角色名",
    "Quality": "品质",
    "HeroClass": "职业定位",
    "Tags": "标签",
    "DamageType": "伤害类型",
    "MaxLv": "最高等级",
    "BasePA60": "60级基础物攻",
    "BaseMA60": "60级基础魔攻",
    "BaseHP60": "60级基础生命",
    "BaseArmor60": "60级基础护甲",
    "WeaponName": "专武名称",
    "WeaponBinding": "专武绑定依据",
    "WeaponSkillId": "专武技能编号",
    "MainAtkFull": "满配主攻击",
    "PAFull": "满配物攻",
    "MAFull": "满配魔攻",
    "HPFull": "满配生命",
    "ArmorFull": "满配护甲",
    "DmgBonus": "总伤害加成",
    "PhysicsDmgBonus": "物理伤害加成",
    "MagicDmgBonus": "魔法伤害加成",
    "NormalAttackDmgBonus": "普攻伤害加成",
    "SkillDmgBonus": "技能伤害加成",
    "UltimateDmgBonus": "必杀伤害加成",
    "SingleTargetDmgBonus": "单体伤害加成",
    "AOEDmgBonus": "群体伤害加成",
    "PenetrationDmgBonus": "穿透伤害加成",
    "CritDmgBonus": "暴伤加成",
    "AttackSpeedBonus": "攻速加成",
    "EnergyPreAddBonus": "初始能量加成",
    "DmgImmune": "减伤免伤",
    "UltimateEffectivePct": "必杀有效倍率估算",
    "ActiveCyclePct": "循环技能倍率估算",
    "AttackSpeedWindowBonus": "攻速窗口收益估算",
    "NormalTriggerBonus": "普攻触发收益估算",
    "BasicAttackLoopBonus": "普攻循环收益估算",
    "SingleKW": "单体关键词数",
    "AoeKW": "群体关键词数",
    "BurstKW": "爆发关键词数",
    "SustainKW": "持续关键词数",
    "SurvivalKW": "生存关键词数",
    "SupportKW": "辅助关键词数",
    "ControlKW": "控制关键词数（首领无效）",
    "EnergyKW": "能量关键词数",
    "CooldownKW": "冷却关键词数",
    "HealPctText": "治疗百分比文本值",
    "ShieldPctText": "护盾百分比文本值",
    "TeamBuffPctText": "团队增益百分比文本值",
    "EnemyDebuffPctText": "敌方减益百分比文本值",
    "TeamKW": "团队关键词数",
    "SkillIds": "技能编号链",
    "EvidenceText": "技能证据文本",
    "Warnings": "风险提示",
    "EmptyEENodes": "空能效节点数",
    "BossScore": "首领单体分",
    "AoeScore": "群体清场分",
    "BurstScore": "爆发分",
    "SustainScore": "持续输出分",
    "SurvivalScore": "生存分",
    "SupportScore": "辅助分",
    "ControlScore": "非首领控制分",
    "OverallScore": "综合分",
    "PeerAvgOverall": "同组综合均值",
    "PeerAvg": "同组均值",
    "PeerDeltaPct": "同组偏差",
    "BenchmarkName": "标杆角色",
    "BenchmarkRole": "标杆口径",
    "BenchmarkCompareMode": "标杆比较模式",
    "BenchmarkOverall": "标杆综合分",
    "BenchmarkDeltaPct": "标杆偏差",
    "BenchmarkFlag": "标杆判断",
    "BenchmarkConfidence": "标杆置信度",
    "BenchmarkSource": "标杆来源",
    "Flag": "判断标记",
    "PrimaryHighScore": "最高子项",
    "Suggestion": "调整建议",
    "KeyEvidence": "关键证据",
    "CheckType": "检查类型",
    "BenchmarkKey": "标杆分组",
    "TopDimension": "最高维度",
    "ModelObservation": "模型观察",
    "CalibrationAction": "校准动作建议",
}


def display_headers(keys: list[str]) -> list[str]:
    missing = [k for k in keys if k not in FIELD_LABELS]
    if missing:
        raise KeyError(f"缺少中文表头映射：{', '.join(missing)}")
    return [FIELD_LABELS[k] for k in keys]


def peer_delta_style(value: Any) -> tuple[PatternFill | None, Font | None]:
    if not isinstance(value, (int, float)):
        return None, None
    for _, lower, upper, fill_color, font_color in PEER_DELTA_COLOR_RULES:
        if lower is not None and value < lower:
            continue
        if upper is not None and value > upper:
            continue
        return PatternFill("solid", fgColor=fill_color), Font(name=BODY_FONT_NAME, color=font_color)
    return None, None


def style_peer_delta_column(sheet: Any, internal_headers: list[str]) -> None:
    if "PeerDeltaPct" not in internal_headers:
        return
    col = internal_headers.index("PeerDeltaPct") + 1
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, col)
        cell.number_format = "0.0%"
        fill, font = peer_delta_style(cell.value)
        if fill and font:
            cell.fill = fill
            cell.font = font


def rows_by_var(path: Path, sheet: str | None = None) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet or wb.sheetnames[0]]
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    fields: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(header):
        name = f"__col{i + 1}" if h is None or str(h).startswith("##") else str(h)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        fields.append(name)
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if any(v is not None for v in r):
            rows.append({fields[i]: r[i] if i < len(r) else None for i in range(len(fields))})
    return rows, fields


def parse_ids(value: Any) -> list[int]:
    if value is None:
        return []
    if isinstance(value, int):
        return [value]
    return [int(p) for p in re.split(r"[;,，；\s]+", str(value).strip()) if p.isdigit()]


def tokens(value: Any) -> list[str]:
    if value is None:
        return []
    return [x.strip() for x in re.split(r"[;；,，\s]+", str(value)) if x.strip()]


def parse_nums(value: Any) -> list[float]:
    vals: list[float] = []
    for t in tokens(value):
        try:
            vals.append(float(t.replace("%", "")))
        except ValueError:
            pass
    return vals


def parse_pct_values(value: Any) -> list[float]:
    vals: list[float] = []
    for t in tokens(value):
        if "%" in t:
            try:
                vals.append(float(t.replace("%", "")))
            except ValueError:
                pass
    return vals


def parse_attr_list(value: Any) -> dict[str, float]:
    out: dict[str, float] = {}
    if not value:
        return out
    for part in str(value).split(";"):
        bits = [x.strip() for x in part.split(",")]
        if len(bits) >= 2:
            try:
                out[bits[0]] = float(bits[1])
            except ValueError:
                pass
    return out


def count_any(text: str, words: list[str]) -> int:
    return sum(text.count(w) for w in words)


def contextual_pct_buckets_for_skill(skill: dict[str, Any]) -> dict[str, float]:
    buckets = {"heal": 0.0, "shield": 0.0, "team": 0.0, "enemy": 0.0}
    desc = re.sub(r"<[^>]+>", "", str(skill.get("Describe") or ""))
    params = tokens(skill.get("DescParam"))
    if not desc or not params:
        return buckets
    windows = desc.split("%s")
    candidates = {"heal": [], "shield": [], "team": [], "enemy": []}
    has_enemy_cap = False

    def has(text: str, words: list[str]) -> bool:
        return any(w in text for w in words)

    for i, raw in enumerate(params):
        if "%" not in raw:
            continue
        try:
            pct = float(raw.replace("%", ""))
        except ValueError:
            continue
        before = windows[i][-48:] if i < len(windows) else ""
        after = windows[i + 1][:48] if i + 1 < len(windows) else ""
        ctx = before + after

        heal_ctx = has(ctx, ["治疗", "回复生命", "恢复生命", "治疗效果", "治疗效率"])
        shield_ctx = has(ctx, ["护盾", "屏障"])
        team_ctx = has(ctx, ["友方", "队友", "全队", "我方全体", "全体友方", "全体队友", "我方"])
        target_ctx = has(ctx, ["目标", "敌人", "敌方", "所有敌"])
        damage_amp_ctx = has(ctx, ["受到的伤害提升", "受到的物理伤害提升", "受到的魔法伤害提升", "易伤"])
        defense_down_ctx = target_ctx and has(ctx, ["防御", "护甲"]) and has(ctx, ["降低", "下降", "减少"])

        if heal_ctx:
            candidates["heal"].append(pct)
        if shield_ctx and has(ctx, ["最大生命", "生命", "强度", "获得", "拥有", "生成", "屏障"]):
            candidates["shield"].append(pct)
        if team_ctx and has(ctx, ["伤害提升", "攻击提升", "攻击速度", "攻速", "暴击", "防御提升", "护甲提升", "减伤", "治疗效率", "技能伤害"]):
            candidates["team"].append(pct)
        if damage_amp_ctx or defense_down_ctx:
            candidates["enemy"].append(pct)
            if "最高叠加" in ctx or "叠加至" in ctx:
                has_enemy_cap = True

    buckets["heal"] = sum(candidates["heal"]) / 100
    buckets["shield"] = sum(candidates["shield"]) / 100
    buckets["team"] = sum(candidates["team"]) / 100
    if has_enemy_cap and candidates["enemy"]:
        buckets["enemy"] = max(candidates["enemy"]) / 100
    else:
        buckets["enemy"] = sum(candidates["enemy"]) / 100
    return buckets


def avg(vals: list[float]) -> float:
    return sum(vals) / len(vals) if vals else 1.0


def build_features(config_root: Path) -> list[dict[str, Any]]:
    heroes, _ = rows_by_var(config_root / "Hero" / "Hero.xlsx", "Hero")
    attrs, _ = rows_by_var(config_root / "Hero" / "HeroAttribute.xlsx", "HeroAttribute")
    stars, _ = rows_by_var(config_root / "Hero" / "HeroStarAttrUp.xlsx", "HeroStar")
    skills, _ = rows_by_var(config_root / "Skill" / "Skill.xlsx", "Skill")
    eeattrs, _ = rows_by_var(config_root / "Hero" / "HeroEnergyEfficiencyAttr.xlsx", "HeroEnergyEfficiencyAttr")
    weapons, _ = rows_by_var(config_root / "ExclusiveWeapon" / "ExclWeapon.xlsx", "Eqpt")

    skill_map = {int(x["SkillId"]): x for x in skills if isinstance(x.get("SkillId"), int)}

    attr_by_hero: dict[int, dict[int, dict[str, Any]]] = {}
    for a in attrs:
        aid = a.get("AttributeID")
        if isinstance(aid, int):
            attr_by_hero.setdefault(aid // 1000, {})[aid % 1000] = a

    star_by_hero: dict[int, list[dict[str, Any]]] = {}
    for s in stars:
        if isinstance(s.get("HeroId"), int):
            star_by_hero.setdefault(s["HeroId"], []).append(s)
    for star_rows in star_by_hero.values():
        star_rows.sort(key=lambda r: r.get("Star") or 0)

    ee_attr_by_id = {r.get("ID"): r for r in eeattrs if isinstance(r.get("ID"), int)}
    weapon_by_name = {
        str(w.get("__col8") or "").strip(): w
        for w in weapons
        if str(w.get("IsValid")).lower() == "true" and w.get("__col8")
    }

    slot_defs = [
        ("Melee", "SkillMelee"),
        ("Power", "SkillsPower"),
        ("Passive1", "SkillsSTEnhancementPassive1"),
        ("Passive2", "SkillsSTEnhancementPassive2"),
        ("Passive3", "SkillsSTEnhancementPassive3"),
        ("Enh1", "SkillsSTEnhancement1"),
        ("Enh2", "SkillsSTEnhancement2"),
        ("Enh3", "SkillsSTEnhancement3"),
        ("Enh4", "SkillsSTEnhancement4"),
    ]
    star_coef_fields = ["HpCoef", "PhysicsAttackCoef", "MagicAttackCoef", "ArmorCoef"]
    star_extra_fields = [
        "AttackSpeed",
        "EnergyPreAdd",
        "EnergyPreAddBonus",
        "Crit",
        "CritDmgBonus",
        "Heal",
        "EffectHit",
        "EffectResist",
        "Miss",
        "Hit",
        "BlockLevel",
        "BreakLevel",
        "BlockStrengthLevel",
        "CritLevel",
        "CritDmgBonusLevel",
        "DmgBonus",
        "DmgImmune",
        "PhysicsDmgBonus",
        "MagicDmgBonus",
        "NormalAttackDmgBonus",
        "SkillDmgBonus",
        "UltimateDmgBonus",
        "SingleTargetDmgBonus",
        "PenetrationDmgBonus",
        "AOEDmgBonus",
    ]
    ee_bonus_fields = [
        "HpBonus",
        "PhysicsAttackBonus",
        "MagicAttackBonus",
        "ArmorBonus",
        "Crit",
        "CritDmgBonus",
        "AttackSpeedBonus",
        "EnergyPreAddBonus",
        "DmgBonus",
        "DmgImmune",
        "PhysicsDmgBonus",
        "MagicDmgBonus",
        "NormalAttackDmgBonus",
        "SkillDmgBonus",
        "UltimateDmgBonus",
        "SingleTargetDmgBonus",
        "PenetrationDmgBonus",
        "AOEDmgBonus",
        "Heal",
    ]

    def skill_text(sid: int | None) -> str:
        if not sid:
            return ""
        sk = skill_map.get(sid, {})
        return " ".join(
            str(sk.get(k) or "")
            for k in ["Name", "Describe", "DescParam", "Introduction", "SkillType", "AutoSkillType", "SkillData"]
        ).replace("\n", " ")

    def latest_skill_ids(hero: dict[str, Any]) -> dict[str, int]:
        out: dict[str, int] = {}
        for label, field in slot_defs:
            ids = parse_ids(hero.get(field))
            if ids:
                out[label] = ids[-1]
        return out

    def effective_pct_for_skill(sid: int | None) -> float:
        sk = skill_map.get(sid, {}) if sid else {}
        desc = str(sk.get("Describe") or "")
        param_text = tokens(sk.get("DescParam"))
        pct_pairs: list[tuple[int, float]] = []
        for i, raw in enumerate(param_text):
            if "%" not in raw:
                continue
            try:
                pct_pairs.append((i, float(raw.replace("%", ""))))
            except ValueError:
                pass
        if not pct_pairs:
            return 0.0
        max_idx, max_pct = max(pct_pairs, key=lambda x: x[1])
        clean_desc = re.sub(r"<[^>]+>", "", desc)
        windows = clean_desc.split("%s")
        count_units = ["枚", "发", "段", "次", "把", "支", "道", "层", "颗", "个", "束", "轮", "波", "下"]
        for i in range(max_idx - 1, -1, -1):
            if i >= len(param_text) or i + 1 >= len(windows):
                continue
            raw = param_text[i]
            if "%" in raw:
                continue
            try:
                count = float(raw)
            except ValueError:
                continue
            if not (1 < count <= 20 and abs(count - int(count)) < 1e-9):
                continue
            after_placeholder = windows[i + 1][:8]
            if any(unit in after_placeholder for unit in count_units) and max_pct < 1000:
                return max_pct * count
        return max_pct

    def basic_attack_loop_bonus(slot_ids: dict[str, int]) -> tuple[float, float, float]:
        speed_bonus = 0.0
        trigger_bonus = 0.0
        for sid in slot_ids.values():
            sk = skill_map.get(sid, {})
            desc = re.sub(r"<[^>]+>", "", str(sk.get("Describe") or ""))
            params = tokens(sk.get("DescParam"))
            windows = desc.split("%s")

            def num_at(i: int) -> float | None:
                if i < 0 or i >= len(params):
                    return None
                try:
                    return float(params[i].replace("%", ""))
                except ValueError:
                    return None

            for i, raw in enumerate(params):
                if "%" not in raw or i >= len(windows):
                    continue
                before = windows[i][-12:]
                if "攻击速度" not in before and "攻速" not in before:
                    continue
                pct = num_at(i)
                if pct is None:
                    continue
                duration = None
                for j in range(i + 1, min(i + 4, len(params))):
                    before_j = windows[j][-8:] if j < len(windows) else ""
                    after_j = windows[j + 1][:8] if j + 1 < len(windows) else ""
                    if "秒" in after_j and "持续" in before_j + after_j:
                        duration = num_at(j)
                        break
                uptime = min(duration / 20, 1.0) if duration else (0.75 if "生命比例高于" in desc else 1.0)
                speed_bonus += pct / 100 * uptime

            if "普攻" not in desc:
                continue
            interval = None
            extra_count = None
            extra_index = None
            for i in range(len(params)):
                before = windows[i][-12:] if i < len(windows) else ""
                after = windows[i + 1][:12] if i + 1 < len(windows) else ""
                val = num_at(i)
                if val is None:
                    continue
                if interval is None and "次普攻" in after:
                    interval = val
                if interval and extra_count is None and ("额外发射" in before + after or "额外" in before) and "次" in after:
                    extra_count = val
                    extra_index = i
            if interval and extra_count and interval > 0 and extra_index is not None:
                damage_pct = 0.0
                duration = None
                cooldown = None
                for j in range(extra_index + 1, len(params)):
                    before_j = windows[j][-8:] if j < len(windows) else ""
                    after_j = windows[j + 1][:8] if j + 1 < len(windows) else ""
                    val = num_at(j)
                    if val is None:
                        continue
                    if "%" in params[j]:
                        damage_pct += val
                    elif "秒" in after_j and "持续" in before_j + after_j:
                        duration = val
                    elif "秒" in after_j and "冷却" in before_j + after_j:
                        cooldown = val
                uptime = min(duration / cooldown, 1.0) if duration and cooldown else 1.0
                trigger_bonus += min(extra_count / interval, 2.0) * damage_pct / 100 * uptime

            if "普攻有" in desc and "概率无视" in desc and "防御" in desc:
                pct_values = [num_at(i) for i, raw in enumerate(params) if "%" in raw]
                pct_values = [v for v in pct_values if v is not None]
                if len(pct_values) >= 2:
                    trigger_bonus += (pct_values[0] / 100) * (pct_values[1] / 100) * 0.6
            if "普攻有" in desc and "概率回复" in desc and "能量" in desc:
                pct_values = [num_at(i) for i, raw in enumerate(params) if "%" in raw]
                if pct_values and pct_values[0] is not None:
                    trigger_bonus += pct_values[0] / 100 * 0.08
        speed_bonus = min(speed_bonus, 1.2)
        trigger_bonus = min(trigger_bonus, 0.8)
        return round(speed_bonus, 4), round(trigger_bonus, 4), round(min(speed_bonus + trigger_bonus, 1.5), 4)

    def star_full(hid: int) -> tuple[dict[str, float], dict[str, float], list[str]]:
        star_rows = star_by_hero.get(hid, [])
        coef = {k: 1.0 for k in star_coef_fields}
        extras: dict[str, float] = {}
        mismatch: list[str] = []
        for r in star_rows:
            desc = str(r.get("StarDesc") or "")
            if desc:
                if "PhysicsAttack" in desc and (r.get("PhysicsAttackCoef") in (None, 0, 1)) and r.get("MagicAttackCoef") not in (None, 0, 1):
                    mismatch.append("升星描述写物攻但实际加魔攻系数")
                if "MagicAttack" in desc and (r.get("MagicAttackCoef") in (None, 0, 1)) and r.get("PhysicsAttackCoef") not in (None, 0, 1):
                    mismatch.append("升星描述写魔攻但实际加物攻系数")
            for k in star_coef_fields:
                v = r.get(k)
                if isinstance(v, (int, float)) and v:
                    coef[k] = max(coef[k], float(v))
            for k in star_extra_fields:
                v = r.get(k)
                if isinstance(v, (int, float)) and v:
                    extras[k] = extras.get(k, 0.0) + float(v)
        return coef, extras, sorted(set(mismatch))

    def energy_full(hero: dict[str, Any]) -> tuple[dict[str, float], int]:
        ids = parse_ids(hero.get("EnergyEfficiencyAttrList"))
        bonus: dict[str, float] = {}
        empty = 0
        for aid in ids:
            r = ee_attr_by_id.get(aid, {})
            had = False
            for k in ee_bonus_fields:
                v = r.get(k)
                if isinstance(v, (int, float)) and v:
                    bonus[k] = bonus.get(k, 0.0) + float(v)
                    had = True
            if not had:
                empty += 1
        return bonus, empty

    def weapon_full(name: Any) -> dict[str, Any] | None:
        w = weapon_by_name.get(str(name).strip())
        if not w:
            return None
        ids = parse_ids(w.get("SkillList"))
        sid = ids[-1] if ids else None
        return {
            "WeaponName": w.get("Name"),
            "PA60": (w.get("PhysicsAttack") or 0) + (w.get("PhysicsAttackGrow") or 0) * 59,
            "MA60": (w.get("MagicAttack") or 0) + (w.get("MagicAttackGrow") or 0) * 59,
            "HP60": (w.get("Hp") or 0) + (w.get("HpGrow") or 0) * 59,
            "Armor60": (w.get("Armor") or 0) + (w.get("ArmorGrow") or 0) * 59,
            "Attr": parse_attr_list(w.get("AttrNameList")),
            "SkillId": sid,
            "SkillText": skill_text(sid),
            "BindingEvidence": "ExclWeapon说明列角色名匹配；RelatedHero为空",
        }

    features: list[dict[str, Any]] = []
    for hero in heroes:
        hid = hero.get("HeroId")
        if not isinstance(hid, int) or hid < 1001:
            continue

        lvmax = max(attr_by_hero.get(hid, {}) or [0])
        a = attr_by_hero.get(hid, {}).get(lvmax, {})
        coef, star_extra, mismatch = star_full(hid)
        ee, empty_ee = energy_full(hero)
        w = weapon_full(hero.get("Name"))
        slot_ids = latest_skill_ids(hero)
        text_parts = [skill_text(sid) for sid in slot_ids.values()]
        if w:
            text_parts.append(w["SkillText"])
        text = " ".join(text_parts)
        tag = str(hero.get("HeroTag") or "")
        skill_effective = {k: effective_pct_for_skill(v) for k, v in slot_ids.items()}
        ultimate_pct = skill_effective.get("Power", 0.0)
        active_cycle_pct = sum(skill_effective.get(k, 0.0) for k in ["Passive1", "Passive2", "Passive3", "Enh1", "Enh2", "Enh3", "Enh4"])
        speed_window_bonus, normal_trigger_bonus, basic_attack_loop = basic_attack_loop_bonus(slot_ids)
        w_attr = w["Attr"] if w else {}

        def sattr(k: str) -> float:
            return float(star_extra.get(k, 0.0) + ee.get(k, 0.0) + w_attr.get(k, 0.0))

        base_pa = a.get("PhysicsAttack") or 0
        base_ma = a.get("MagicAttack") or 0
        base_hp = a.get("Hp") or 0
        base_armor = a.get("Armor") or 0
        pa = (base_pa * coef["PhysicsAttackCoef"] + (w["PA60"] if w else 0)) * (
            1 + star_extra.get("PhysicsAttackBonus", 0) + ee.get("PhysicsAttackBonus", 0) + w_attr.get("PhysicsAttackBonus", 0)
        )
        ma = (base_ma * coef["MagicAttackCoef"] + (w["MA60"] if w else 0)) * (
            1 + star_extra.get("MagicAttackBonus", 0) + ee.get("MagicAttackBonus", 0) + w_attr.get("MagicAttackBonus", 0)
        )
        hp = (base_hp * coef["HpCoef"] + (w["HP60"] if w else 0)) * (
            1 + star_extra.get("HpBonus", 0) + ee.get("HpBonus", 0) + w_attr.get("HpBonus", 0)
        )
        armor = (base_armor * coef["ArmorCoef"] + (w["Armor60"] if w else 0)) * (
            1 + star_extra.get("ArmorBonus", 0) + ee.get("ArmorBonus", 0) + w_attr.get("ArmorBonus", 0)
        )
        if "魔法攻击" in text and "物理攻击" not in text:
            dtype = "MAG"
        elif "物理攻击" in text and "魔法攻击" not in text:
            dtype = "PHY"
        elif "物理攻击" in text and "魔法攻击" in text:
            dtype = "MIX"
        else:
            dtype = "UNK"
        main = max(pa, ma) if dtype in ("MIX", "UNK") else (pa if dtype == "PHY" else ma)

        def kw(words: list[str]) -> int:
            return count_any(tag + " " + text, words)

        pct_buckets = {"heal": 0.0, "shield": 0.0, "team": 0.0, "enemy": 0.0}
        for sid in slot_ids.values():
            for k, v in contextual_pct_buckets_for_skill(skill_map.get(sid, {})).items():
                pct_buckets[k] += v
        if w and w.get("SkillId"):
            for k, v in contextual_pct_buckets_for_skill(skill_map.get(w["SkillId"], {})).items():
                pct_buckets[k] += v
        support_pct = pct_buckets["team"]
        heal_pct = pct_buckets["heal"]
        shield_pct = pct_buckets["shield"]
        enemy_debuff_pct = pct_buckets["enemy"]

        features.append(
            {
                "HeroId": hid,
                "Name": hero.get("Name"),
                "Quality": hero.get("Quality"),
                "HeroClass": hero.get("HeroClass"),
                "Tags": tag,
                "DamageType": dtype,
                "MaxLv": lvmax,
                "BasePA60": base_pa,
                "BaseMA60": base_ma,
                "BaseHP60": base_hp,
                "BaseArmor60": base_armor,
                "WeaponName": w["WeaponName"] if w else "",
                "WeaponBinding": w["BindingEvidence"] if w else "",
                "WeaponSkillId": w["SkillId"] if w else "",
                "MainAtkFull": round(main, 2),
                "PAFull": round(pa, 2),
                "MAFull": round(ma, 2),
                "HPFull": round(hp, 2),
                "ArmorFull": round(armor, 2),
                "DmgBonus": sattr("DmgBonus"),
                "PhysicsDmgBonus": sattr("PhysicsDmgBonus"),
                "MagicDmgBonus": sattr("MagicDmgBonus"),
                "NormalAttackDmgBonus": sattr("NormalAttackDmgBonus"),
                "SkillDmgBonus": sattr("SkillDmgBonus"),
                "UltimateDmgBonus": sattr("UltimateDmgBonus"),
                "SingleTargetDmgBonus": sattr("SingleTargetDmgBonus"),
                "AOEDmgBonus": sattr("AOEDmgBonus"),
                "PenetrationDmgBonus": sattr("PenetrationDmgBonus"),
                "CritDmgBonus": sattr("CritDmgBonus"),
                "AttackSpeedBonus": sattr("AttackSpeedBonus"),
                "EnergyPreAddBonus": sattr("EnergyPreAddBonus"),
                "DmgImmune": sattr("DmgImmune"),
                "UltimateEffectivePct": ultimate_pct,
                "ActiveCyclePct": active_cycle_pct,
                "AttackSpeedWindowBonus": speed_window_bonus,
                "NormalTriggerBonus": normal_trigger_bonus,
                "BasicAttackLoopBonus": basic_attack_loop,
                "SingleKW": kw(["单体", "目标", "最高", "最近", "boss", "Boss", "BOSS"]),
                "AoeKW": kw(["群体", "范围", "全体", "全队", "所有敌", "穿透", "多个", "小范围", "大范围", "周围", "前方所有"]),
                "BurstKW": kw(["爆发", "必杀", "暴击", "暴击伤害", "无视", "额外", "释放必杀", "斩杀", "已损失"]),
                "SustainKW": kw(["持续", "每秒", "普攻", "攻击速度", "冷却", "召唤", "幻象", "每隔", "叠加", "永久"]),
                "SurvivalKW": kw(["护盾", "治疗", "回复", "生命", "减伤", "免疫", "防御", "格挡"]),
                "SupportKW": kw(["友方", "全体", "全队", "队友", "增益", "攻击提升", "伤害提升", "攻速", "能量", "治疗", "护盾"]),
                "ControlKW": kw(["冰冻", "沉默", "眩晕", "控制", "击退", "减速", "打断", "定身", "中毒", "流血"]),
                "EnergyKW": kw(["能量", "回能", "回复怒气", "回复能量", "不耗能"]),
                "CooldownKW": kw(["冷却", "冷却时间降低", "技能冷却"]),
                "HealPctText": heal_pct,
                "ShieldPctText": shield_pct,
                "TeamBuffPctText": support_pct,
                "EnemyDebuffPctText": enemy_debuff_pct,
                "TeamKW": kw(["友方", "队友", "全队", "我方全体", "全体友方", "全体队友", "我方"]),
                "SkillIds": ";".join(f"{k}:{v}" for k, v in slot_ids.items()),
                "EvidenceText": text[:1200],
                "Warnings": ";".join(mismatch),
                "EmptyEENodes": empty_ee,
            }
        )
    return features


def add_scores(features: list[dict[str, Any]]) -> None:
    class_quality_avg: dict[tuple[Any, Any], list[float]] = defaultdict(list)
    class_quality_hp: dict[tuple[Any, Any], list[float]] = defaultdict(list)
    class_quality_armor: dict[tuple[Any, Any], list[float]] = defaultdict(list)
    class_avg: dict[Any, list[float]] = defaultdict(list)
    class_hp: dict[Any, list[float]] = defaultdict(list)
    class_armor: dict[Any, list[float]] = defaultdict(list)

    for f in features:
        key = (f["HeroClass"], f["Quality"])
        class_quality_avg[key].append(float(f["MainAtkFull"]))
        class_quality_hp[key].append(float(f["HPFull"]))
        class_quality_armor[key].append(float(f["ArmorFull"]))
        class_avg[f["HeroClass"]].append(float(f["MainAtkFull"]))
        class_hp[f["HeroClass"]].append(float(f["HPFull"]))
        class_armor[f["HeroClass"]].append(float(f["ArmorFull"]))

    def val(d: dict[str, Any], k: str) -> float:
        v = d.get(k)
        return float(v) if isinstance(v, (int, float)) else 0.0

    for f in features:
        target_atk = avg(class_quality_avg[(f["HeroClass"], f["Quality"])]) or avg(class_avg[f["HeroClass"]])
        target_hp = avg(class_quality_hp[(f["HeroClass"], f["Quality"])]) or avg(class_hp[f["HeroClass"]])
        target_armor = avg(class_quality_armor[(f["HeroClass"], f["Quality"])]) or avg(class_armor[f["HeroClass"]])
        atk_ratio = val(f, "MainAtkFull") / target_atk
        hp_ratio = val(f, "HPFull") / target_hp
        armor_ratio = val(f, "ArmorFull") / target_armor
        enemy = val(f, "EnemyDebuffPctText")
        basic_attack_loop = val(f, "BasicAttackLoopBonus")
        attack_speed_window = val(f, "AttackSpeedWindowBonus")
        normal_trigger = val(f, "NormalTriggerBonus")
        team_buff = val(f, "TeamBuffPctText") if val(f, "TeamKW") > 0 else 0
        heal_team = val(f, "HealPctText") if val(f, "TeamKW") > 0 or f["HeroClass"] in ("Support", "Tank") else 0
        shield_team = val(f, "ShieldPctText") if val(f, "TeamKW") > 0 or f["HeroClass"] in ("Support", "Tank") else 0

        boss = 100 * atk_ratio * (
            1
            + val(f, "DmgBonus")
            + 0.35 * val(f, "PhysicsDmgBonus")
            + 0.35 * val(f, "MagicDmgBonus")
            + 0.25 * val(f, "NormalAttackDmgBonus")
            + 0.35 * val(f, "SkillDmgBonus")
            + 0.60 * val(f, "UltimateDmgBonus")
            + 0.55 * val(f, "SingleTargetDmgBonus")
            + 0.25 * val(f, "PenetrationDmgBonus")
            + 0.20 * val(f, "CritDmgBonus")
            + 0.20 * val(f, "AttackSpeedBonus")
            + min(val(f, "UltimateEffectivePct"), 1800) / 3500
            + 0.35 * basic_attack_loop
            + 0.03 * val(f, "SingleKW")
            + 0.02 * val(f, "BurstKW")
            + 0.15 * enemy
        )
        aoe = 100 * atk_ratio * (
            1
            + val(f, "DmgBonus")
            + 0.35 * val(f, "SkillDmgBonus")
            + 0.35 * val(f, "UltimateDmgBonus")
            + 0.70 * val(f, "AOEDmgBonus")
            + 0.10 * basic_attack_loop
            + 0.08 * val(f, "AoeKW")
            + min(val(f, "UltimateEffectivePct"), 1500) / 5000
            + 0.10 * enemy
        )
        burst = 100 * atk_ratio * (
            1
            + val(f, "DmgBonus")
            + 0.80 * val(f, "UltimateDmgBonus")
            + 0.20 * val(f, "CritDmgBonus")
            + 0.15 * val(f, "AttackSpeedBonus")
            + 0.08 * val(f, "EnergyKW")
            + 0.45 * attack_speed_window
            + 0.25 * normal_trigger
            + min(val(f, "UltimateEffectivePct"), 2000) / 2500
            + 0.04 * val(f, "BurstKW")
        )
        sustain = 100 * atk_ratio * (
            1
            + val(f, "DmgBonus")
            + 0.65 * val(f, "NormalAttackDmgBonus")
            + 0.35 * val(f, "SkillDmgBonus")
            + 0.45 * val(f, "AttackSpeedBonus")
            + 0.25 * val(f, "EnergyPreAddBonus")
            + 0.70 * basic_attack_loop
            + 0.05 * val(f, "SustainKW")
            + 0.04 * val(f, "CooldownKW")
            + min(val(f, "ActiveCyclePct"), 1200) / 6000
        )
        survival = 100 * (
            0.55 * hp_ratio
            + 0.25 * armor_ratio
            + 0.06 * val(f, "SurvivalKW")
            + 0.25 * val(f, "ShieldPctText")
            + 0.20 * val(f, "HealPctText")
            + 0.40 * val(f, "DmgImmune")
        )
        support = min(
            260,
            40
            + 16 * val(f, "TeamKW")
            + 45 * team_buff
            + 30 * enemy
            + 35 * heal_team
            + 35 * shield_team
            + 8 * val(f, "EnergyKW")
            + 6 * val(f, "CooldownKW"),
        )
        if f["HeroClass"] == "DPS":
            support = min(support, 120 + 20 * val(f, "TeamKW") + 15 * enemy)
        if f["HeroClass"] == "Tank":
            support = min(support, 180 + 20 * val(f, "TeamKW") + 20 * enemy)
        control = min(220, 25 + 18 * val(f, "ControlKW") + 18 * enemy)

        if f["HeroClass"] == "DPS":
            overall = 0.30 * boss + 0.20 * aoe + 0.17 * burst + 0.25 * sustain + 0.05 * survival + 0.03 * control
        elif f["HeroClass"] == "Tank":
            overall = 0.52 * survival + 0.15 * control + 0.13 * support + 0.10 * boss + 0.10 * aoe
        else:
            overall = 0.52 * support + 0.15 * survival + 0.13 * control + 0.10 * boss + 0.10 * aoe
        f.update(
            {
                "BossScore": round(boss, 1),
                "AoeScore": round(aoe, 1),
                "BurstScore": round(burst, 1),
                "SustainScore": round(sustain, 1),
                "SurvivalScore": round(survival, 1),
                "SupportScore": round(support, 1),
                "ControlScore": round(control, 1),
                "OverallScore": round(overall, 1),
            }
        )

    avg_overall: dict[tuple[Any, Any], list[float]] = defaultdict(list)
    for f in features:
        avg_overall[(f["HeroClass"], f["Quality"])].append(float(f["OverallScore"]))

    benchmark_features: dict[tuple[Any, Any], dict[str, Any]] = {}
    for key, benchmark in CHARACTER_BENCHMARKS.items():
        matched = next((x for x in features if x.get("HeroId") == benchmark["HeroId"]), None)
        if matched:
            benchmark_features[key] = matched

    def benchmark_for(f: dict[str, Any]) -> tuple[dict[str, Any] | None, dict[str, Any] | None, str]:
        key = (f["HeroClass"], f["Quality"])
        if key in benchmark_features:
            return CHARACTER_BENCHMARKS[key], benchmark_features[key], "同品质比较"
        if isinstance(f.get("Quality"), int) and f["Quality"] < 3:
            sr_key = (f["HeroClass"], 3)
            if sr_key in benchmark_features:
                return CHARACTER_BENCHMARKS[sr_key], benchmark_features[sr_key], "低品质上限检查"
        return None, None, "无标杆"

    for f in features:
        baseline = avg(avg_overall[(f["HeroClass"], f["Quality"])])
        delta = (float(f["OverallScore"]) - baseline) / baseline if baseline else 0
        flag = "偏强" if delta >= 0.22 else ("偏弱" if delta <= -0.18 else "接近均值")
        if f.get("Warnings"):
            flag += ";配置风险"
        f["PeerAvgOverall"] = round(baseline, 1)
        f["PeerDeltaPct"] = round(delta, 3)
        f["Flag"] = flag

        benchmark, benchmark_feature, benchmark_mode = benchmark_for(f)
        if benchmark and benchmark_feature:
            benchmark_overall = float(benchmark_feature["OverallScore"])
            benchmark_delta = (float(f["OverallScore"]) - benchmark_overall) / benchmark_overall if benchmark_overall else 0
            if f.get("HeroId") == benchmark["HeroId"]:
                benchmark_flag = "标杆本体"
            elif benchmark_mode == "低品质上限检查":
                benchmark_flag = "越级高于SR标杆风险" if benchmark_delta >= 0 else "低于SR标杆（符合品质预期）"
            elif benchmark_delta >= 0.15:
                benchmark_flag = "显著高于标杆"
            elif benchmark_delta <= -0.15:
                benchmark_flag = "显著低于标杆"
            else:
                benchmark_flag = "接近标杆"
            f["BenchmarkName"] = benchmark["Name"]
            f["BenchmarkRole"] = "低品质对SR上限" if benchmark_mode == "低品质上限检查" else benchmark["Role"]
            f["BenchmarkCompareMode"] = benchmark_mode
            f["BenchmarkOverall"] = round(benchmark_overall, 1)
            f["BenchmarkDeltaPct"] = round(benchmark_delta, 3)
            f["BenchmarkFlag"] = benchmark_flag
            f["BenchmarkConfidence"] = benchmark["Confidence"]
            f["BenchmarkSource"] = benchmark["Source"]
        else:
            f["BenchmarkName"] = ""
            f["BenchmarkRole"] = "待补标杆"
            f["BenchmarkCompareMode"] = benchmark_mode
            f["BenchmarkOverall"] = ""
            f["BenchmarkDeltaPct"] = ""
            f["BenchmarkFlag"] = "无高置信标杆"
            f["BenchmarkConfidence"] = "无"
            f["BenchmarkSource"] = ""

        if f["HeroClass"] == "DPS":
            candidates = [
                ("Boss", f["BossScore"]),
                ("AOE", f["AoeScore"]),
                ("Burst", f["BurstScore"]),
                ("Sustain", f["SustainScore"]),
                ("Survival", f["SurvivalScore"]),
                ("Control", f["ControlScore"]),
            ]
        elif f["HeroClass"] == "Tank":
            candidates = [
                ("Survival", f["SurvivalScore"]),
                ("Control", f["ControlScore"]),
                ("Support", f["SupportScore"]),
                ("Boss", f["BossScore"]),
                ("AOE", f["AoeScore"]),
            ]
        else:
            candidates = [
                ("Support", f["SupportScore"]),
                ("Survival", f["SurvivalScore"]),
                ("Control", f["ControlScore"]),
                ("Boss", f["BossScore"]),
                ("AOE", f["AoeScore"]),
            ]
        high = max(candidates, key=lambda x: x[1])
        f["PrimaryHighScore"] = f"{high[0]}:{high[1]}"

        suggestions: list[str] = []
        if "偏强" in flag:
            if f["HeroClass"] == "DPS":
                if high[0] == "AOE":
                    suggestions.append("先下调群体覆盖/范围技能/必杀系数 8%-12%")
                elif high[0] == "Boss":
                    suggestions.append("先下调单体乘区、无视防御、易伤或高血锁定收益 8%-12%")
                elif high[0] == "Burst":
                    suggestions.append("先下调必杀伤害/回能/开局能量 8%-12%")
                elif high[0] == "Sustain":
                    suggestions.append("先下调普攻增伤、攻速、冷却或长期叠层 8%-12%")
                elif high[0] == "Survival":
                    suggestions.append("先下调护盾/治疗/低血收益 10%-15%，避免 DPS 越界")
                else:
                    suggestions.append("Boss 不吃控制；仅按小怪/波次/PVPVE价值检查控制覆盖")
            elif f["HeroClass"] == "Tank":
                suggestions.append("先下调护盾/治疗/减伤覆盖率 10%-15%，不要直接砍基础生命")
            else:
                suggestions.append("先下调团队增益覆盖率或幅度 8%-12%，检查是否全队常驻")
        elif "偏弱" in flag:
            if f["HeroClass"] == "DPS":
                suggestions.append("优先补主定位乘区 8%-12%，不要同时补单体和群体")
            elif f["HeroClass"] == "Tank":
                suggestions.append("优先补有效护盾/减伤覆盖或防御转化，避免补输出")
            else:
                suggestions.append("优先补核心辅助价值或覆盖率，避免让辅助输出追上 DPS")
        else:
            suggestions.append("先不改主数值；若实测偏差，按最高子项微调 3%-5%")
        if f["HeroClass"] == "DPS" and f["SurvivalScore"] > 150:
            suggestions.append("DPS 生存分过高，查护盾/治疗/低血收益是否越界")
        if float(f.get("UltimateDmgBonus") or 0) >= 0.8:
            suggestions.append("必杀乘区过高，检查 UltimateDmgBonus 与能效/专武叠加")
        if float(f.get("SkillDmgBonus") or 0) >= 0.8:
            suggestions.append("技能乘区过高，检查 SkillDmgBonus 与技能频率")
        if f.get("Warnings"):
            suggestions.append("先修/确认配置风险，再做强度结论")
        f["Suggestion"] = "；".join(suggestions)


SCORE_DIMENSIONS = [
    ("BossScore", "首领单体分"),
    ("AoeScore", "群体清场分"),
    ("BurstScore", "爆发分"),
    ("SustainScore", "持续输出分"),
    ("SurvivalScore", "生存分"),
    ("SupportScore", "辅助分"),
    ("ControlScore", "非首领控制分"),
]


def benchmark_calibration_rows(features: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    feature_by_id = {f.get("HeroId"): f for f in features}
    grouped: dict[tuple[Any, Any], list[dict[str, Any]]] = defaultdict(list)
    for f in features:
        grouped[(f.get("HeroClass"), f.get("Quality"))].append(f)

    def top_dimension(f: dict[str, Any]) -> str:
        key, label = max(SCORE_DIMENSIONS, key=lambda item: float(f.get(item[0]) or 0))
        return f"{label}:{float(f.get(key) or 0):.1f}"

    def base_row(f: dict[str, Any], check_type: str, action: str, observation: str = "") -> dict[str, Any]:
        key = f"{f.get('HeroClass')}-Q{f.get('Quality')}"
        return {
            "CheckType": check_type,
            "BenchmarkKey": key,
            "HeroId": f.get("HeroId"),
            "Name": f.get("Name"),
            "Quality": f.get("Quality"),
            "HeroClass": f.get("HeroClass"),
            "BenchmarkName": f.get("BenchmarkName"),
            "BenchmarkRole": f.get("BenchmarkRole"),
            "OverallScore": f.get("OverallScore"),
            "BenchmarkOverall": f.get("BenchmarkOverall"),
            "BenchmarkDeltaPct": f.get("BenchmarkDeltaPct"),
            "PeerDeltaPct": f.get("PeerDeltaPct"),
            "BossScore": f.get("BossScore"),
            "AoeScore": f.get("AoeScore"),
            "BurstScore": f.get("BurstScore"),
            "SustainScore": f.get("SustainScore"),
            "SurvivalScore": f.get("SurvivalScore"),
            "SupportScore": f.get("SupportScore"),
            "ControlScore": f.get("ControlScore"),
            "TopDimension": top_dimension(f),
            "ModelObservation": observation,
            "CalibrationAction": action,
            "BenchmarkConfidence": f.get("BenchmarkConfidence"),
            "Warnings": f.get("Warnings"),
        }

    def benchmark_observation(f: dict[str, Any]) -> tuple[str, str]:
        dims = {k: float(f.get(k) or 0) for k, _ in SCORE_DIMENSIONS}
        top = top_dimension(f)
        if f.get("HeroClass") == "DPS":
            dps_core = max(dims["BossScore"], dims["AoeScore"], dims["BurstScore"], dims["SustainScore"])
            if dims["SurvivalScore"] > dps_core * 0.85:
                return f"标杆最高维度为{top}，但生存分接近输出核心项", "检查 DPS 生存权重和护盾/治疗关键词，避免标杆被生存项抬高"
            return f"标杆最高维度为{top}，输出核心项主导", "暂不调权重；后续用实战 Boss/清场数据校准输出子项占比"
        if f.get("HeroClass") == "Tank":
            if dims["SurvivalScore"] < max(dims["BossScore"], dims["AoeScore"]):
                return f"标杆最高维度为{top}，输出项高于生存项", "优先检查坦克公式，避免输出项支配坦克综合分"
            return f"标杆最高维度为{top}，生存/保护项主导", "暂不调权重；后续用承伤、护盾覆盖、死亡率校准"
        if dims["SupportScore"] < max(dims["BossScore"], dims["AoeScore"]):
            return f"标杆最高维度为{top}，输出项高于辅助项", "优先检查辅助公式，避免输出项支配奶妈/辅助综合分"
        return f"标杆最高维度为{top}，辅助/生存项主导", "暂不调权重；后续用治疗有效值、护盾覆盖和团队DPS提升校准"

    for key, benchmark in CHARACTER_BENCHMARKS.items():
        f = feature_by_id.get(benchmark["HeroId"])
        if not f:
            rows.append(
                {
                    "CheckType": "标杆缺失",
                    "BenchmarkKey": f"{key[0]}-Q{key[1]}",
                    "HeroId": benchmark["HeroId"],
                    "Name": benchmark["Name"],
                    "Quality": key[1],
                    "HeroClass": key[0],
                    "BenchmarkName": benchmark["Name"],
                    "BenchmarkRole": benchmark["Role"],
                    "OverallScore": "",
                    "BenchmarkOverall": "",
                    "BenchmarkDeltaPct": "",
                    "PeerDeltaPct": "",
                    "BossScore": "",
                    "AoeScore": "",
                    "BurstScore": "",
                    "SustainScore": "",
                    "SurvivalScore": "",
                    "SupportScore": "",
                    "ControlScore": "",
                    "TopDimension": "",
                    "ModelObservation": "配置中未找到该标杆角色",
                    "CalibrationAction": "先确认 Hero.xlsx 角色名和 HeroId，再做标杆校准",
                    "BenchmarkConfidence": benchmark["Confidence"],
                    "Warnings": "",
                }
            )
            continue
        observation, action = benchmark_observation(f)
        rows.append(base_row(f, "标杆本体", action, observation))

    for key, group in sorted(grouped.items(), key=lambda item: (str(item[0][0]), item[0][1] or 0)):
        cls, quality = key
        if key not in CHARACTER_BENCHMARKS and not (isinstance(quality, int) and quality < 3 and (cls, 3) in CHARACTER_BENCHMARKS):
            top = max(group, key=lambda x: float(x.get("OverallScore") or 0))
            rows.append(
                base_row(
                    top,
                    "待补标杆",
                    "需要用户指定该分组的高置信标杆，否则只能看同组均值",
                    f"{cls}-Q{quality} 当前无高置信标杆；本行展示该组综合分最高角色",
                )
            )

    for f in features:
        if f.get("HeroId") in {b["HeroId"] for b in CHARACTER_BENCHMARKS.values()}:
            continue
        delta = f.get("BenchmarkDeltaPct")
        if not isinstance(delta, (int, float)):
            continue
        if f.get("BenchmarkCompareMode") == "低品质上限检查":
            if delta < 0:
                continue
            rows.append(
                base_row(
                    f,
                    "低品质越级风险",
                    "优先检查低品质角色面板、技能倍率或权重；原则上低品质不应超过高品质标杆",
                    f"该低品质角色综合分高于同职业 SR 标杆 {f.get('BenchmarkName')}",
                )
            )
        elif delta >= 0.15:
            rows.append(
                base_row(
                    f,
                    "高于标杆",
                    "检查该角色最高子项是否应该高于标杆；若无明确定位成本，优先下调最高子项",
                    f"相对标杆 {f.get('BenchmarkName')} 高 {delta:.1%}",
                )
            )
        elif delta <= -0.15:
            rows.append(
                base_row(
                    f,
                    "低于标杆",
                    "检查是否符合定位差异；若同定位但缺乏成本补偿，优先补主定位乘区",
                    f"相对标杆 {f.get('BenchmarkName')} 低 {abs(delta):.1%}",
                )
            )
    return rows


def write_outputs(features: list[dict[str, Any]], output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "vgame_character_fullmax_quant_v2.xlsx"
    md_path = output_dir / "vgame_character_fullmax_quant_v2_summary.md"

    wb = Workbook()
    scores = wb.active
    scores.title = "ScoresSnapshot"
    raw = wb.create_sheet("RawFeatures")
    assumptions = wb.create_sheet("Assumptions")
    calibration = wb.create_sheet("标杆校准检查")
    suggestions = wb.create_sheet("Suggestions")

    score_headers = [
        "RankInClass",
        "HeroId",
        "Name",
        "Quality",
        "HeroClass",
        "Tags",
        "WeaponName",
        "MainAtkFull",
        "HPFull",
        "ArmorFull",
        "BossScore",
        "AoeScore",
        "BurstScore",
        "SustainScore",
        "SurvivalScore",
        "SupportScore",
        "ControlScore",
        "OverallScore",
        "PeerAvgOverall",
        "PeerDeltaPct",
        "BenchmarkName",
        "BenchmarkRole",
        "BenchmarkOverall",
        "BenchmarkDeltaPct",
        "BenchmarkFlag",
        "BenchmarkConfidence",
        "Flag",
        "PrimaryHighScore",
        "Suggestion",
        "Warnings",
    ]
    scores.append(display_headers(score_headers))
    for cls in sorted({f["HeroClass"] for f in features}):
        rows = sorted([x for x in features if x["HeroClass"] == cls], key=lambda x: -float(x["OverallScore"]))
        for rank, f in enumerate(rows, 1):
            scores.append([rank] + [f.get(h, "") for h in score_headers[1:]])

    raw_headers = list(features[0].keys()) if features else []
    raw.append(display_headers(raw_headers))
    for f in features:
        raw.append([f.get(k, "") for k in raw_headers])

    assumptions.append(["项目", "说明"])
    for r in [
        ["版本", "v2.1，修正非辅助角色 SupportScore 被自增伤顶满的问题，并修正治疗/护盾/团队增益/敌方减益的百分比上下文归因。"],
        ["口径", "60级 + 满星 + 满能效图；具名专武按 ExclWeapon 说明列角色名匹配纳入，通用专武不强绑角色。"],
        ["分数含义", "配置可见启发式指数，不等于 runtime 实战 DPS；同职业同品质内做 PeerDelta。"],
        ["标杆校准检查", "新增 sheet：标杆校准检查，用于检查标杆本体、同品质相对标杆离群、低品质越级和待补标杆分组。"],
        ["偏强阈值", "同职业同品质综合分 +22% 以上"],
        ["偏弱阈值", "同职业同品质综合分 -18% 以下"],
        ["同组偏差颜色", "红=+22%以上偏强；橙=+10%到+22%偏高观察；绿=-10%到+10%接近均值；浅蓝=-18%到-10%偏低观察；蓝=-18%以下偏弱。"],
        ["首领控制规则", "Vgame 首领不吃控制；首领单体分不计控制关键词，非首领控制分仅代表小怪/波次/PVPVE等场景价值。"],
        ["多段倍率识别", "只把伤害百分比之前、且占位符后紧跟枚/发/段/次等数量单位的参数当作段数；冷却秒数、持续秒数、概率次数不再放大倍率。"],
        ["普攻每秒倍率", "技能描述中“每秒造成X%，Y发/秒”按每秒总倍率X%理解；枚/发/霰弹/发射频率不直接放大每秒总伤害，除非SkillData明确每枚独立结算。"],
        ["普攻循环折算", "攻速持续窗口、每N次普攻额外子弹、普攻概率无视防御、普攻回能会折成普攻循环收益；主要进入持续输出，部分进入首领单体和爆发。"],
        ["百分比上下文归因", "治疗、护盾、团队增益、敌方易伤/降防按单个技能描述的%s占位符前后文归类；自增伤、自攻速、伤害倍率、持续秒数不再整包累加。"],
        ["DPS权重", "Boss 30%, AOE 20%, Burst 17%, Sustain 25%, Survival 5%, Control 3%"],
        ["Tank权重", "Survival 52%, Control 15%, Support 13%, Boss 10%, AOE 10%"],
        ["Support权重", "Support 52%, Survival 15%, Control 13%, Boss 10%, AOE 10%"],
        ["高置信标杆", "输出SSR=陈单单；输出SR=克莉丝汀（用户称克里斯汀）；坦克SSR=塔拉夏；坦克SR=大暑；奶妈SR=陈双双。"],
        ["标杆偏差", "标杆偏差 = (角色综合分 - 同职业同品质标杆综合分) / 标杆综合分；没有同品质标杆时标记为待补。"],
        ["低品质规则", "低品质角色对高品质标杆只做上限检查；低于高品质标杆属于符合品质预期，不视为低于风险。"],
        ["限制", "未解析 SkillData bytes；多段命中、Buff覆盖、召唤继承、Boss免疫需实测复核。"],
    ]:
        assumptions.append(r)

    calibration_headers = [
        "CheckType",
        "BenchmarkKey",
        "HeroId",
        "Name",
        "Quality",
        "HeroClass",
        "BenchmarkName",
        "BenchmarkRole",
        "OverallScore",
        "BenchmarkOverall",
        "BenchmarkDeltaPct",
        "PeerDeltaPct",
        "BossScore",
        "AoeScore",
        "BurstScore",
        "SustainScore",
        "SurvivalScore",
        "SupportScore",
        "ControlScore",
        "TopDimension",
        "ModelObservation",
        "CalibrationAction",
        "BenchmarkConfidence",
        "Warnings",
    ]
    calibration.append(display_headers(calibration_headers))
    calibration_rows = benchmark_calibration_rows(features)
    for row in calibration_rows:
        calibration.append([row.get(k, "") for k in calibration_headers])

    suggestion_headers = [
        "HeroId",
        "Name",
        "Quality",
        "HeroClass",
        "OverallScore",
        "PeerAvg",
        "PeerDeltaPct",
        "BenchmarkName",
        "BenchmarkOverall",
        "BenchmarkDeltaPct",
        "BenchmarkFlag",
        "Flag",
        "PrimaryHighScore",
        "Suggestion",
        "KeyEvidence",
    ]
    suggestions.append(display_headers(suggestion_headers))
    for f in sorted(features, key=lambda x: (not str(x["Flag"]).startswith("偏强"), -abs(float(x["PeerDeltaPct"])), -float(x["OverallScore"]))):
        evidence = (
            f"Tags={f['Tags']} | W={f['WeaponName'] or '-'} | UltPct={f['UltimateEffectivePct']} | "
            f"SkillBonus={float(f.get('SkillDmgBonus') or 0):.2f} | UltBonus={float(f.get('UltimateDmgBonus') or 0):.2f} | "
            f"TeamKW={f['TeamKW']} | Warnings={f.get('Warnings') or '-'}"
        )
        suggestions.append(
            [
                f["HeroId"],
                f["Name"],
                f["Quality"],
                f["HeroClass"],
                f["OverallScore"],
                f["PeerAvgOverall"],
                f["PeerDeltaPct"],
                f["BenchmarkName"],
                f["BenchmarkOverall"],
                f["BenchmarkDeltaPct"],
                f["BenchmarkFlag"],
                f["Flag"],
                f["PrimaryHighScore"],
                f["Suggestion"],
                evidence,
            ]
        )

    header_font = Font(name=HEADER_FONT_NAME, bold=True, color="FFFFFF")
    body_font = Font(name=BODY_FONT_NAME, color="000000")
    table_names = {
        "ScoresSnapshot": "ScoresSnapshotTable",
        "RawFeatures": "RawFeaturesTable",
        "Assumptions": "AssumptionsTable",
        "标杆校准检查": "BenchmarkCalibrationTable",
        "Suggestions": "SuggestionsTable",
    }
    for sh in [scores, raw, assumptions, calibration, suggestions]:
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = sh.dimensions
        for c in sh[1]:
            c.font = header_font
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row != 1:
                    cell.font = body_font
        for col in range(1, sh.max_column + 1):
            width = min(max(len(str(sh.cell(r, col).value or "")[:70]) for r in range(1, min(sh.max_row, 80) + 1)) + 2, 42)
            sh.column_dimensions[get_column_letter(col)].width = width
        table = Table(displayName=table_names[sh.title], ref=sh.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sh.add_table(table)
    score_one_decimal = [
        "MainAtkFull",
        "HPFull",
        "ArmorFull",
        "BossScore",
        "AoeScore",
        "BurstScore",
        "SustainScore",
        "SurvivalScore",
        "SupportScore",
        "ControlScore",
        "OverallScore",
        "PeerAvgOverall",
        "BenchmarkOverall",
    ]
    for row in scores.iter_rows(min_row=2):
        for key in score_one_decimal:
            row[score_headers.index(key)].number_format = "0.0"
        row[score_headers.index("BenchmarkDeltaPct")].number_format = "0.0%"
    style_peer_delta_column(scores, score_headers)
    style_peer_delta_column(raw, raw_headers)
    calibration_one_decimal = [
        "OverallScore",
        "BenchmarkOverall",
        "BossScore",
        "AoeScore",
        "BurstScore",
        "SustainScore",
        "SurvivalScore",
        "SupportScore",
        "ControlScore",
    ]
    for row in calibration.iter_rows(min_row=2):
        for key in calibration_one_decimal:
            row[calibration_headers.index(key)].number_format = "0.0"
        row[calibration_headers.index("BenchmarkDeltaPct")].number_format = "0.0%"
    style_peer_delta_column(calibration, calibration_headers)
    suggestion_one_decimal = ["OverallScore", "PeerAvg", "BenchmarkOverall"]
    for row in suggestions.iter_rows(min_row=2):
        for key in suggestion_one_decimal:
            row[suggestion_headers.index(key)].number_format = "0.0"
        row[suggestion_headers.index("BenchmarkDeltaPct")].number_format = "0.0%"
    style_peer_delta_column(suggestions, suggestion_headers)
    wb.save(xlsx_path)

    def pct_text(value: Any) -> str:
        return f"{float(value):.1%}" if isinstance(value, (int, float)) else "-"

    lines = [
        "# Vgame 全角色全满量化 v2.1",
        "",
        "口径：60级、满星、满能效图、最高档技能；具名专武按 ExclWeapon 说明列角色名匹配纳入。v2.1 修正了 DPS 自增伤误计辅助和百分比整包串读的问题。",
        "标杆：输出SSR陈单单、输出SR克莉丝汀、坦克SSR塔拉夏、坦克SR大暑、奶妈SR陈双双；R卡只检查是否超过同职业SR标杆。",
        "",
    ]
    for cls in ["DPS", "Tank", "Support"]:
        rows = sorted([f for f in features if f["HeroClass"] == cls], key=lambda x: -float(x["OverallScore"]))
        lines.append(f"## {cls} 排名")
        lines.append("|排名|角色|品质|综合分|同组偏差|标杆|标杆偏差|判断|主项|建议|")
        lines.append("|---:|---|---:|---:|---:|---|---:|---|---|---|")
        for rank, f in enumerate(rows, 1):
            lines.append(
                f"|{rank}|{f['HeroId']} {f['Name']}|{f['Quality']}|{float(f['OverallScore']):.1f}|{float(f['PeerDeltaPct']):.1%}|"
                f"{f['BenchmarkName'] or '-'}|{pct_text(f['BenchmarkDeltaPct'])}|{f['Flag']}|{f['PrimaryHighScore']}|{str(f['Suggestion'])[:90]}|"
            )
        lines.append("")
    lines.append("## 优先处理")
    for f in sorted([x for x in features if str(x["Flag"]).startswith("偏强")], key=lambda x: -float(x["PeerDeltaPct"]))[:10]:
        lines.append(
            f"- 偏强：{f['HeroId']} {f['Name']} {f['HeroClass']} Q{f['Quality']}，综合 {float(f['OverallScore']):.1f}，"
            f"同组偏差 {float(f['PeerDeltaPct']):.1%}，主项 {f['PrimaryHighScore']}。{f['Suggestion']}"
        )
    for f in sorted([x for x in features if str(x["Flag"]).startswith("偏弱")], key=lambda x: float(x["PeerDeltaPct"]))[:10]:
        lines.append(
            f"- 偏弱：{f['HeroId']} {f['Name']} {f['HeroClass']} Q{f['Quality']}，综合 {float(f['OverallScore']):.1f}，"
            f"同组偏差 {float(f['PeerDeltaPct']):.1%}，主项 {f['PrimaryHighScore']}。{f['Suggestion']}"
        )
    lines.append("")
    lines.append("## 标杆校准检查")
    lines.append("详见 Excel 的 `标杆校准检查` sheet；该页用于检查标杆本体、同品质相对标杆离群、低品质越级和待补标杆分组。")
    priority_types = {"高于标杆", "低品质越级风险", "待补标杆"}
    for row in [x for x in calibration_rows if x.get("CheckType") in priority_types][:12]:
        delta = pct_text(row.get("BenchmarkDeltaPct"))
        lines.append(
            f"- {row['CheckType']}：{row['HeroId']} {row['Name']} {row['HeroClass']} Q{row['Quality']}，"
            f"标杆={row.get('BenchmarkName') or '-'}，标杆偏差={delta}，{row['CalibrationAction']}"
        )
    md_path.write_text("\n".join(lines), encoding="utf-8")
    return xlsx_path, md_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Vgame full-max character quantification workbook.")
    parser.add_argument("--config-root", type=Path, default=DEFAULT_CONFIG_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()
    features = build_features(args.config_root)
    add_scores(features)
    xlsx_path, md_path = write_outputs(features, args.output_dir)
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    main()
