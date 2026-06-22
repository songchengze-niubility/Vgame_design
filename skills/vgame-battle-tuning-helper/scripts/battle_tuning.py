"""
Vgame Battle Tuning Helper
============================
从 Vgame基本战斗框架.xlsx 读取怪物/玩家数据，计算 DPS/TTK/存活/难度评估。

Usage:
    python battle_tuning.py --monster-level 50
    python battle_tuning.py --stage 30
    python battle_tuning.py --stage 50 --monster-level 100
    python battle_tuning.py --stage 30 --hp-coeff 1.5
    python battle_tuning.py --mode tower --floor 50
    python battle_tuning.py --mode infinite --wave 20
"""

import argparse
import os
import sys

BATTLE_XLSX_PATH = r"D:\数值文档\数值文档\Vgame基本战斗框架.xlsx"
DEF_CONSTANT = 24000
BASE_BLOCK_REDUCTION = 0.5
DMG_REDUCTION_CAP = 0.75
HIT_FLOOR = 0.10
DMG_FLOOR_RATIO = 0.05


def load_workbook_cached():
    """加载 Excel（只读模式）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(BATTLE_XLSX_PATH):
        print(f"[ERROR] 未找到战斗框架文件: {BATTLE_XLSX_PATH}")
        sys.exit(1)

    return load_workbook(BATTLE_XLSX_PATH, read_only=True, data_only=True)


def get_monster_stats(wb, level: int) -> dict:
    """
    从 Sheet '怪物标准模型' 读取指定等级的怪物属性。
    Row = 11 + level (Level 1 = Row 12, Level 500 = Row 511)
    """
    ws = wb["怪物标准模型"]
    row = 11 + level

    def val(col):
        v = ws.cell(row=row, column=col).value
        return float(v) if v is not None else 0.0

    return {
        "level": level,
        "final_hp": val(26),       # Z: 生命(最终)
        "final_atk": val(27),      # AA: 攻击(最终)
        "final_def": val(28),      # AB: 防御(最终)
        "dmg_reduction": val(29),  # AC: 减伤率
        "ehp": val(30),            # AD: 有效HP
        "ttk": val(32),            # AF: TTK(秒)
        "party_coeff": val(33),    # AG: 人数系数
        "atk_coeff": val(34),      # AH: 攻击系数修正
        "boss_hp_coeff": val(35),  # AI: BOSS生命系数修正
        "mob_hp_coeff": val(36),   # AJ: 小怪生命系数修正
        "mob_atk_coeff": val(37),  # AK: 小怪攻击系数修正
        "boss_hp": val(38),        # AL: BOSS生命
        "boss_atk": val(39),       # AM: BOSS攻击
        "boss_def": val(40),       # AN: BOSS防御
        "mob_hp": val(41),         # AO: 小怪生命
        "mob_atk": val(42),        # AP: 小怪攻击
    }


def get_player_stats(wb, stage: int) -> dict:
    """
    从 Sheet '玩家战力曲线' 读取指定关卡的玩家预期属性。
    Row = 11 + stage (Stage 1 = Row 12, Stage 105 = Row 116)
    """
    ws = wb["玩家战力曲线"]
    row = 11 + stage

    def val(col):
        v = ws.cell(row=row, column=col).value
        return float(v) if v is not None else 0.0

    return {
        "stage": stage,
        "player_level": val(1),     # A: 玩家等级
        "chapter": val(3),          # C: 章节
        "stage_in_chapter": val(4), # D: 关卡
        "final_def": val(15),       # O: 防御(最终)
        "dmg_reduction": val(16),   # P: 减伤率
        "final_atk": val(20),       # T: 攻击(最终)
        "normal_coeff": val(21),    # U: 普攻系数
        "skill_coeff": val(22),     # V: 技能系数
        "ult_coeff": val(23),       # W: 必杀系数
        "crit_rate": val(33),       # AG: 暴击率
        "crit_dmg": val(34),        # AH: 暴击伤害
        "dmg_bonus": val(35),       # AI: 增伤倍率
        "total_dps": val(37),       # AK: 总单位时间伤害
        "expected_boss_ttk": val(38),  # AL: 预计bossTTK
        "boss_hp": val(40),         # AN: boss生命
        "monster_level": val(41),   # AO: 怪物等级
    }


def calc_damage(atk: float, skill_coeff: float, defender_def: float, ignore_def: float = 0) -> float:
    """计算单次伤害"""
    def_effective = defender_def * (1 - ignore_def)
    penetration = 1 - def_effective / (def_effective + DEF_CONSTANT)
    base_dmg = atk * skill_coeff * penetration
    return max(base_dmg, atk * skill_coeff * DMG_FLOOR_RATIO)


def calc_dps(atk: float, defender_def: float, crit_rate: float, crit_dmg: float,
             dmg_bonus: float, normal_coeff: float = 1.0, skill_coeff: float = 1.0,
             ult_coeff: float = 1.0, ignore_def: float = 0) -> float:
    """计算 DPS"""
    # 加权平均系数: 普攻1.0/s + 技能0.4/s + 必杀0.55/s = 1.95/s
    avg_coeff = (normal_coeff * 1.0 + skill_coeff * 0.4 + ult_coeff * 0.55) / 1.95
    base_dmg_per_hit = calc_damage(atk, avg_coeff, defender_def, ignore_def)
    crit_multiplier = 1 + crit_rate * (crit_dmg - 1)
    dmg_multiplier = 1 + dmg_bonus
    return base_dmg_per_hit * crit_multiplier * dmg_multiplier * 1.95


def calc_ttk(hp: float, dmg_reduction: float, dps: float) -> float:
    """计算击杀时间"""
    if dps <= 0:
        return float('inf')
    effective_hp = hp / (1 - min(dmg_reduction, DMG_REDUCTION_CAP)) if dmg_reduction < 1 else hp * 100
    return effective_hp / dps


def assess_difficulty(ttk: float, expected_time: float, survival_time: float) -> str:
    """评估难度"""
    if survival_time < ttk:
        return "[X] 打不过（存活时间 < 击杀时间）"
    ratio = ttk / expected_time if expected_time > 0 else 1.0
    if ratio < 0.5:
        return "[碾压] 过于简单"
    elif ratio < 0.8:
        return "[简单] 偏简单"
    elif ratio < 1.2:
        return "[适中] 难度合理"
    elif ratio < 1.5:
        return "[偏难] 有挑战"
    else:
        return "[过难] 可能超时/卡关"


def infinite_challenge_level(wave: int) -> int:
    """无限挑战怪物等级公式: level = 27 + 1000*(n-1)/((n-1)+38)"""
    if wave <= 1:
        return 27
    n = wave
    level = 27 + 1000 * (n - 1) / ((n - 1) + 38)
    return min(500, int(level))


def print_monster_info(stats: dict, hp_coeff: float = 1.0, atk_coeff: float = 1.0):
    """打印怪物信息"""
    print(f"\n{'='*50}")
    print(f"  怪物等级: {stats['level']}")
    print(f"{'='*50}")
    print(f"\n基准属性:")
    print(f"  最终HP:  {stats['final_hp']:,.0f}")
    print(f"  最终ATK: {stats['final_atk']:,.0f}")
    print(f"  最终DEF: {stats['final_def']:,.0f}")
    print(f"  减伤率:  {stats['dmg_reduction']:.2%}")
    print(f"  有效HP:  {stats['ehp']:,.0f}")
    print(f"\nBoss 属性:")
    print(f"  Boss HP:  {stats['boss_hp'] * hp_coeff:,.0f}" + (f" (×{hp_coeff})" if hp_coeff != 1.0 else ""))
    print(f"  Boss ATK: {stats['boss_atk'] * atk_coeff:,.0f}" + (f" (×{atk_coeff})" if atk_coeff != 1.0 else ""))
    print(f"  Boss DEF: {stats['boss_def']:,.0f}")
    print(f"\n小怪属性:")
    print(f"  小怪 HP:  {stats['mob_hp']:,.0f}")
    print(f"  小怪 ATK: {stats['mob_atk']:,.0f}")
    print(f"\n系数:")
    print(f"  Boss HP系数: {stats['boss_hp_coeff']:.2f}")
    print(f"  小怪HP系数:  {stats['mob_hp_coeff']:.2f}")
    print(f"  攻击系数:    {stats['atk_coeff']:.2f}")
    print(f"  人数系数:    {stats['party_coeff']:.2f}")


def print_player_info(stats: dict):
    """打印玩家信息"""
    print(f"\n{'='*50}")
    print(f"  玩家 - 关卡{int(stats['stage'])} (Ch{int(stats['chapter'])}-{int(stats['stage_in_chapter'])})")
    print(f"{'='*50}")
    print(f"  玩家等级:  {int(stats['player_level'])}")
    print(f"  怪物等级:  {int(stats['monster_level'])}")
    print(f"  最终ATK:   {stats['final_atk']:,.0f}")
    print(f"  最终DEF:   {stats['final_def']:,.0f}")
    print(f"  减伤率:    {stats['dmg_reduction']:.2%}")
    print(f"  暴击率:    {stats['crit_rate']:.2%}")
    print(f"  暴击伤害:  {stats['crit_dmg']:.2f}")
    print(f"  增伤倍率:  {stats['dmg_bonus']:.2%}")
    print(f"  总DPS:     {stats['total_dps']:,.0f}")
    print(f"  预计BossTTK: {stats['expected_boss_ttk']:.1f}s")
    print(f"  Boss HP:   {stats['boss_hp']:,.0f}")


def print_combat_analysis(player: dict, monster: dict, hp_coeff: float = 1.0, atk_coeff: float = 1.0):
    """打印战斗分析"""
    # 玩家打Boss: 优先使用玩家曲线里的 Boss HP（已经过玩法难度线系数修正）
    if player.get("boss_hp", 0) > 0:
        boss_hp = player["boss_hp"] * hp_coeff
    else:
        boss_hp = monster["boss_hp"] * hp_coeff

    boss_def = monster["boss_def"]
    boss_atk = monster["boss_atk"] * atk_coeff
    boss_dmg_reduction = monster["dmg_reduction"]

    player_dps = player["total_dps"] if player["total_dps"] > 0 else calc_dps(
        player["final_atk"], boss_def, player["crit_rate"],
        player["crit_dmg"], player["dmg_bonus"],
        player.get("normal_coeff", 1.0), player.get("skill_coeff", 1.0), player.get("ult_coeff", 1.0)
    )

    ttk = calc_ttk(boss_hp, boss_dmg_reduction, player_dps)

    # Boss打玩家（简化：Boss DPS = Boss ATK × 0.5系数 × 1.0/s）
    boss_dps_to_player = calc_damage(boss_atk, 0.5, player["final_def"]) * 1.0
    player_ehp = player["final_def"] * 10 / (1 - min(player["dmg_reduction"], DMG_REDUCTION_CAP)) if player["dmg_reduction"] < 1 else player["final_def"] * 100
    # 粗估玩家HP（ATK:HP = 3:10）
    player_hp_est = player["final_atk"] / 3 * 10
    survival_time = player_hp_est / boss_dps_to_player if boss_dps_to_player > 0 else float('inf')

    expected_time = player["expected_boss_ttk"] if player["expected_boss_ttk"] > 0 else 60.0
    difficulty = assess_difficulty(ttk, expected_time, survival_time)

    print(f"\n{'='*50}")
    print(f"  战斗分析")
    print(f"{'='*50}")
    print(f"\n玩家 → Boss:")
    print(f"  玩家 DPS:       {player_dps:,.0f}")
    print(f"  Boss 有效HP:    {boss_hp / (1 - min(boss_dmg_reduction, 0.75)):,.0f}")
    print(f"  预计 TTK:       {ttk:.1f}s")
    print(f"  预期 TTK:       {expected_time:.1f}s")
    print(f"  TTK/预期 比值:  {ttk/expected_time:.2f}")
    print(f"\nBoss → 玩家:")
    print(f"  Boss DPS(估):   {boss_dps_to_player:,.0f}")
    print(f"  玩家 HP(估):    {player_hp_est:,.0f}")
    print(f"  玩家存活时间:   {survival_time:.1f}s")
    print(f"\n难度评估: {difficulty}")

    if hp_coeff != 1.0 or atk_coeff != 1.0:
        print(f"\n  [应用系数] HP×{hp_coeff}, ATK×{atk_coeff}")


def main():
    parser = argparse.ArgumentParser(description="Vgame Battle Tuning Helper")
    parser.add_argument("--monster-level", type=int, default=0, help="查询怪物等级 (1-500)")
    parser.add_argument("--stage", type=int, default=0, help="查询主线关卡 (1-105)")
    parser.add_argument("--hp-coeff", type=float, default=1.0, help="怪物HP系数覆盖")
    parser.add_argument("--atk-coeff", type=float, default=1.0, help="怪物ATK系数覆盖")
    parser.add_argument("--mode", type=str, default="", help="玩法模式: tower, infinite, mythic")
    parser.add_argument("--floor", type=int, default=0, help="层数/波数 (配合 --mode)")
    parser.add_argument("--wave", type=int, default=0, help="无限挑战波数 (配合 --mode infinite)")
    args = parser.parse_args()

    wb = load_workbook_cached()

    # 模式路由
    monster_level = args.monster_level
    stage = args.stage

    if args.mode == "infinite":
        wave = args.wave or args.floor
        if wave > 0:
            monster_level = infinite_challenge_level(wave)
            print(f"[INFO] 无限挑战 Wave {wave} → 怪物等级 {monster_level}")

    if args.mode == "tower" and args.floor > 0:
        # 爬塔简单近似: level ≈ 20 + floor * 1.8
        monster_level = min(500, int(20 + args.floor * 1.8))
        print(f"[INFO] 爬塔 Floor {args.floor} → 怪物等级 {monster_level} (近似)")

    if args.mode == "mythic" and args.floor > 0:
        # 大秘境: level ≈ 40 + floor * 7.2 (25层覆盖40-220)
        monster_level = min(500, int(40 + args.floor * 7.2))
        print(f"[INFO] 大秘境 Layer {args.floor} → 怪物等级 {monster_level} (近似)")

    # 查询怪物
    if monster_level > 0:
        monster_level = min(500, max(1, monster_level))
        monster = get_monster_stats(wb, monster_level)
        print_monster_info(monster, args.hp_coeff, args.atk_coeff)

    # 查询玩家
    if stage > 0:
        stage = min(105, max(1, stage))
        player = get_player_stats(wb, stage)
        print_player_info(player)

        # 如果同时有怪物等级，做战斗分析
        if monster_level > 0:
            print_combat_analysis(player, monster, args.hp_coeff, args.atk_coeff)
        elif player["monster_level"] > 0:
            # 用玩家曲线自带的怪物等级
            ml = int(player["monster_level"])
            monster = get_monster_stats(wb, ml)
            print_monster_info(monster, args.hp_coeff, args.atk_coeff)
            print_combat_analysis(player, monster, args.hp_coeff, args.atk_coeff)

    # 只有怪物没有玩家
    if monster_level > 0 and stage == 0:
        print(f"\n[提示] 加 --stage N 可对比玩家在第N关时打这个怪物的战斗预期。")

    wb.close()


if __name__ == "__main__":
    main()
