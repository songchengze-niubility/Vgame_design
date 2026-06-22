"""
Vgame Player Progression Simulator
===================================
模拟玩家在 N 天内的资源积累和养成进度。

Usage:
    python simulate_progression.py --days 30 --scenario f2p_active
    python simulate_progression.py --days 60 --scenario f2p_active,light_spender --compare
    python simulate_progression.py --days 30 --scenario f2p_active --override "mainline_gold_multiplier=1.2"
    python simulate_progression.py --days 30 --scenario f2p_active --from-excel
"""

import argparse
import json
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple


# ========== Excel 自动读取 ==========

ECONOMY_XLSX_PATH = r"D:\数值文档\数值文档\Vgame经济框架.xlsx"


def load_from_excel() -> Tuple[dict, dict, dict, dict, dict, dict]:
    """
    从 Vgame经济框架.xlsx 自动读取最新数据。
    返回: (CHARACTER_LEVEL_EXP, CHARACTER_LEVEL_GOLD, WEAPON_LEVEL_EXP, WEAPON_LEVEL_GOLD, MAINLINE_REWARDS, MAINLINE_STAGE_COUNT)

    Excel 结构约定（已验证）:
    - Sheet "角色升级": Row 7-66, B=等级, E=累计经验, G=累计金币
    - Sheet "专武": Row 7-66, B=等级, E=累计经验(SSR), G=累计金币
    - Sheet "主线副本": Row 6-65, A=章节, B=关卡, G=首通金币量, H=经验道具类型名, I=经验数量
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[WARNING] openpyxl 未安装，使用内置默认值。安装: pip install openpyxl")
        return None, None, None, None, None, None

    if not os.path.exists(ECONOMY_XLSX_PATH):
        print(f"[WARNING] 未找到经济框架文件: {ECONOMY_XLSX_PATH}，使用内置默认值。")
        return None, None, None, None, None, None

    print(f"[INFO] 从 Excel 读取最新数据: {ECONOMY_XLSX_PATH}")
    wb = load_workbook(ECONOMY_XLSX_PATH, read_only=True, data_only=True)

    # === 角色升级 ===
    ws = wb["角色升级"]
    char_exp = {}
    char_gold = {}
    for row in range(7, 67):  # Row 7-66, Level 1-60
        level = ws.cell(row=row, column=2).value  # Col B
        cum_exp = ws.cell(row=row, column=5).value  # Col E
        cum_gold = ws.cell(row=row, column=7).value  # Col G
        if level is not None and cum_exp is not None:
            level = int(level)
            # 每5级取一个节点用于模拟
            if level % 5 == 0:
                char_exp[level] = int(cum_exp)
                char_gold[level] = int(cum_gold) if cum_gold else int(cum_exp * 0.5)

    # === 专武 ===
    ws = wb["专武"]
    weap_exp = {}
    weap_gold = {}
    for row in range(7, 67):  # Row 7-66, Level 1-60
        level = ws.cell(row=row, column=2).value  # Col B
        cum_exp = ws.cell(row=row, column=5).value  # Col E
        cum_gold = ws.cell(row=row, column=7).value  # Col G
        if level is not None and cum_exp is not None:
            level = int(level)
            if level % 5 == 0 or level in (10, 20, 30, 40, 50, 60):
                weap_exp[level] = int(cum_exp)
                weap_gold[level] = int(cum_gold) if cum_gold else int(cum_exp * 0.1)

    # === 主线副本 ===
    ws = wb["主线副本"]
    # 按章节统计: 每章的普通金币、Boss金币、经验单价、经验数量普通、经验数量Boss、钻石
    chapter_stages = {}  # {chapter: [(gold, exp_value, exp_qty, diamond), ...]}
    for row in range(6, 66):  # Row 6-65
        chapter = ws.cell(row=row, column=1).value  # Col A
        stage = ws.cell(row=row, column=2).value  # Col B
        gold = ws.cell(row=row, column=7).value  # Col G: 首通金币量
        exp_type_name = ws.cell(row=row, column=8).value  # Col H: 经验道具类型名
        exp_qty = ws.cell(row=row, column=9).value  # Col I: 经验数量
        diamond_qty = ws.cell(row=row, column=5).value  # Col E: 钻石数量

        if chapter is None or gold is None:
            continue

        chapter = int(chapter)
        gold = int(gold)
        exp_qty = int(exp_qty) if exp_qty else 0
        diamond_qty = int(diamond_qty) if diamond_qty else 20

        # 判断经验类型: "角色经验低"=500/个, "角色经验中"=2000/个, "角色经验高"=10000/个
        exp_unit = 2000  # 默认中
        if exp_type_name:
            exp_type_str = str(exp_type_name)
            if "低" in exp_type_str:
                exp_unit = 500
            elif "高" in exp_type_str:
                exp_unit = 10000
            else:
                exp_unit = 2000

        if chapter not in chapter_stages:
            chapter_stages[chapter] = []
        chapter_stages[chapter].append((gold, exp_unit, exp_qty, diamond_qty))

    # 从每章数据中计算: 普通金币(众数), Boss金币(最后一关), 经验单价, 普通数量(众数), Boss数量(最后一关), 钻石
    mainline_rewards = {}
    mainline_stage_count = {}
    for ch, stages in sorted(chapter_stages.items()):
        mainline_stage_count[ch] = len(stages)
        if not stages:
            continue

        # Boss 是最后一关
        boss_gold = stages[-1][0]
        boss_exp_qty = stages[-1][2]
        diamond = stages[0][3]  # 钻石每关一样
        exp_unit = stages[0][1]  # 经验单价每章一样

        # 普通关金币和经验数量取非Boss关的值
        if len(stages) > 1:
            normal_gold = stages[0][0]
            normal_exp_qty = stages[0][2]
        else:
            normal_gold = boss_gold
            normal_exp_qty = boss_exp_qty

        mainline_rewards[ch] = (normal_gold, boss_gold, exp_unit, normal_exp_qty, boss_exp_qty, diamond)

    wb.close()

    print(f"[INFO] 读取完成: 角色{len(char_exp)}个节点, 专武{len(weap_exp)}个节点, 主线{sum(mainline_stage_count.values())}关")
    return char_exp, char_gold, weap_exp, weap_gold, mainline_rewards, mainline_stage_count


# ========== 内置默认参数（当 Excel 不可用时使用） ==========

# 角色升级 EXP 需求（累计）- 每5级一个节点
CHARACTER_LEVEL_EXP = {
    5: 1500, 10: 5000, 15: 12500, 20: 30000, 25: 60000, 30: 105000,
    35: 170000, 40: 280000, 45: 420000, 50: 630000, 55: 1100000, 60: 1922500
}

# 角色升级金币需求（累计，约为 EXP 的 0.5 倍）
CHARACTER_LEVEL_GOLD = {
    5: 750, 10: 2500, 15: 6250, 20: 15000, 25: 30000, 30: 52500,
    35: 85000, 40: 140000, 45: 210000, 50: 315000, 55: 550000, 60: 961250
}

# 专武升级 EXP 需求（累计，SSR）
WEAPON_LEVEL_EXP = {
    5: 5000, 10: 15000, 20: 60000, 30: 210000, 40: 560000, 50: 1260000, 60: 2634500
}

# 专武升级金币需求
WEAPON_LEVEL_GOLD = {
    5: 500, 10: 1500, 20: 6000, 30: 21000, 40: 56000, 50: 126000, 60: 263450
}

# 主线首通奖励
MAINLINE_REWARDS = {
    # chapter: (gold_normal, gold_boss, exp_type, exp_qty_normal, exp_qty_boss, diamond_per_stage)
    0: (5000, 10000, 500, 6, 10, 20),      # 5关, 经验低(500)
    1: (10000, 20000, 2000, 6, 10, 20),     # 10关, 经验中(2000)
    2: (15000, 30000, 2000, 8, 12, 20),     # 15关, 经验中
    3: (20000, 40000, 2000, 12, 20, 20),    # 15关, 经验中
    4: (25000, 50000, 2000, 16, 24, 20),    # 15关, 经验中
}

MAINLINE_STAGE_COUNT = {0: 5, 1: 10, 2: 15, 3: 15, 4: 15}


@dataclass
class ResourceState:
    """资源状态"""
    gold: int = 0
    diamond: int = 0
    character_exp: int = 0  # 以 EXP 值计
    weapon_exp: int = 0
    stamina: int = 0
    gacha_tickets_normal: int = 0
    gacha_tickets_limited: int = 0
    weapon_tickets: int = 0


@dataclass
class ProgressState:
    """进度状态"""
    day: int = 0
    mainline_stage: int = 0  # 已通关的总关卡数 (0-60)
    character_level: int = 1
    character_exp_used: int = 0
    character_gold_used: int = 0
    weapon_level: int = 1
    weapon_exp_used: int = 0
    weapon_gold_used: int = 0
    gacha_pulls: int = 0
    ssr_characters: int = 1  # 初始送1个


@dataclass
class DailyReport:
    """每日报告"""
    day: int
    income: Dict[str, int] = field(default_factory=dict)
    expense: Dict[str, int] = field(default_factory=dict)
    totals: Dict[str, int] = field(default_factory=dict)
    milestones: List[str] = field(default_factory=list)
    bottlenecks: List[str] = field(default_factory=list)


@dataclass
class ScenarioParams:
    """场景参数"""
    name: str = "f2p_active"
    daily_stamina: int = 180          # 每日总体力
    stamina_purchase: int = 0         # 每日购买体力次数
    stamina_per_purchase: int = 60    # 每次购买体力量
    monthly_card_diamond: int = 0     # 月卡每日钻石
    battle_pass_daily: int = 0        # 通行证每日额外资源（钻石等价）
    daily_task_gold: int = 20000      # 每日任务金币
    daily_task_diamond: int = 30      # 每日任务钻石
    daily_task_exp: int = 4000        # 每日任务角色经验
    weekly_task_diamond: int = 150    # 每周任务钻石（均摊到日）
    sign_in_daily_gold: int = 5000    # 签到日均金币
    sign_in_daily_diamond: int = 20   # 签到日均钻石
    resource_dungeon_gold: int = 30000   # 金币本日收入
    resource_dungeon_char_exp: int = 12000  # 角色经验本日收入（EXP值）
    resource_dungeon_weap_exp: int = 8000   # 专武经验本日收入（EXP值）
    mainline_stages_per_day: int = 4  # 每日主线推进关卡数
    gacha_strategy: str = "save_for_pity"
    stamina_waste_ratio: float = 0.0  # 体力浪费比例
    # 覆盖乘数
    mainline_gold_multiplier: float = 1.0
    daily_income_multiplier: float = 1.0


def get_scenario(name: str) -> ScenarioParams:
    """获取预设场景"""
    scenarios = {
        "f2p_active": ScenarioParams(
            name="f2p_active",
            daily_stamina=180, stamina_purchase=0,
            monthly_card_diamond=0, battle_pass_daily=0,
            daily_task_gold=20000, daily_task_diamond=30, daily_task_exp=4000,
            weekly_task_diamond=150, sign_in_daily_gold=5000, sign_in_daily_diamond=20,
            resource_dungeon_gold=30000, resource_dungeon_char_exp=12000,
            resource_dungeon_weap_exp=8000, mainline_stages_per_day=4,
        ),
        "f2p_casual": ScenarioParams(
            name="f2p_casual",
            daily_stamina=180, stamina_purchase=0,
            monthly_card_diamond=0, battle_pass_daily=0,
            daily_task_gold=14000, daily_task_diamond=20, daily_task_exp=2800,
            weekly_task_diamond=100, sign_in_daily_gold=5000, sign_in_daily_diamond=20,
            resource_dungeon_gold=20000, resource_dungeon_char_exp=8000,
            resource_dungeon_weap_exp=5000, mainline_stages_per_day=3,
            stamina_waste_ratio=0.3,
        ),
        "light_spender": ScenarioParams(
            name="light_spender",
            daily_stamina=180, stamina_purchase=1, stamina_per_purchase=60,
            monthly_card_diamond=90, battle_pass_daily=0,
            daily_task_gold=20000, daily_task_diamond=30, daily_task_exp=4000,
            weekly_task_diamond=150, sign_in_daily_gold=5000, sign_in_daily_diamond=20,
            resource_dungeon_gold=40000, resource_dungeon_char_exp=16000,
            resource_dungeon_weap_exp=10000, mainline_stages_per_day=5,
        ),
        "medium_spender": ScenarioParams(
            name="medium_spender",
            daily_stamina=180, stamina_purchase=2, stamina_per_purchase=60,
            monthly_card_diamond=90, battle_pass_daily=30,
            daily_task_gold=20000, daily_task_diamond=30, daily_task_exp=4000,
            weekly_task_diamond=150, sign_in_daily_gold=5000, sign_in_daily_diamond=20,
            resource_dungeon_gold=50000, resource_dungeon_char_exp=20000,
            resource_dungeon_weap_exp=14000, mainline_stages_per_day=5,
        ),
    }
    return scenarios.get(name, scenarios["f2p_active"])


def get_mainline_reward(stage_index: int) -> Dict[str, int]:
    """获取某关的首通奖励"""
    # stage_index 从 0 开始 (0-59)
    cumulative = 0
    for chapter, count in MAINLINE_STAGE_COUNT.items():
        if stage_index < cumulative + count:
            local_index = stage_index - cumulative
            gold_n, gold_b, exp_type, exp_n, exp_b, diamond = MAINLINE_REWARDS[chapter]
            is_boss = (local_index == count - 1)
            gold = gold_b if is_boss else gold_n
            exp_qty = exp_b if is_boss else exp_n
            return {
                "gold": gold,
                "diamond": diamond,
                "character_exp": exp_type * exp_qty,
            }
        cumulative += count
    return {"gold": 0, "diamond": 0, "character_exp": 0}


def get_level_for_exp(exp_used: int, level_table: dict) -> int:
    """根据已使用经验计算等级"""
    level = 1
    for lv, required_exp in sorted(level_table.items()):
        if exp_used >= required_exp:
            level = lv
        else:
            break
    return level


def simulate(params: ScenarioParams, days: int) -> List[DailyReport]:
    """运行模拟"""
    resources = ResourceState()
    progress = ProgressState()
    reports = []

    for day in range(1, days + 1):
        progress.day = day
        income = {}
        expense = {}
        milestones = []
        bottlenecks = []

        # === 每日收入 ===

        # 1. 每日任务 + 签到
        daily_gold = int((params.daily_task_gold + params.sign_in_daily_gold) * params.daily_income_multiplier)
        daily_diamond = int((params.daily_task_diamond + params.sign_in_daily_diamond +
                           params.monthly_card_diamond + params.battle_pass_daily +
                           params.weekly_task_diamond / 7) * params.daily_income_multiplier)
        daily_char_exp = int(params.daily_task_exp * params.daily_income_multiplier)

        income["gold_daily"] = daily_gold
        income["diamond_daily"] = daily_diamond
        income["char_exp_daily"] = daily_char_exp

        resources.gold += daily_gold
        resources.diamond += daily_diamond
        resources.character_exp += daily_char_exp

        # 2. 资源副本
        dungeon_gold = int(params.resource_dungeon_gold * params.daily_income_multiplier)
        dungeon_char_exp = int(params.resource_dungeon_char_exp * params.daily_income_multiplier)
        dungeon_weap_exp = int(params.resource_dungeon_weap_exp * params.daily_income_multiplier)

        income["gold_dungeon"] = dungeon_gold
        income["char_exp_dungeon"] = dungeon_char_exp
        income["weap_exp_dungeon"] = dungeon_weap_exp

        resources.gold += dungeon_gold
        resources.character_exp += dungeon_char_exp
        resources.weapon_exp += dungeon_weap_exp

        # 3. 主线首通（一次性）
        stages_today = min(params.mainline_stages_per_day, 60 - progress.mainline_stage)
        mainline_gold_today = 0
        mainline_diamond_today = 0
        mainline_exp_today = 0

        for i in range(stages_today):
            stage_idx = progress.mainline_stage + i
            reward = get_mainline_reward(stage_idx)
            mainline_gold_today += int(reward["gold"] * params.mainline_gold_multiplier)
            mainline_diamond_today += reward["diamond"]
            mainline_exp_today += reward["character_exp"]

        if stages_today > 0:
            progress.mainline_stage += stages_today
            resources.gold += mainline_gold_today
            resources.diamond += mainline_diamond_today
            resources.character_exp += mainline_exp_today
            income["gold_mainline"] = mainline_gold_today
            income["diamond_mainline"] = mainline_diamond_today
            income["char_exp_mainline"] = mainline_exp_today

        # === 每日消耗（自动养成） ===

        # 优先升级主角色（每天尝试升到资源够的最高级）
        for lv, required_exp in sorted(CHARACTER_LEVEL_EXP.items()):
            if lv <= progress.character_level:
                continue
            exp_needed = required_exp - progress.character_exp_used
            gold_needed = CHARACTER_LEVEL_GOLD[lv] - progress.character_gold_used
            if resources.character_exp >= exp_needed and resources.gold >= gold_needed:
                resources.character_exp -= exp_needed
                resources.gold -= gold_needed
                progress.character_exp_used = required_exp
                progress.character_gold_used = CHARACTER_LEVEL_GOLD[lv]
                progress.character_level = lv
                milestones.append(f"角色达到 Lv{lv}")
                expense["char_level_exp"] = expense.get("char_level_exp", 0) + exp_needed
                expense["char_level_gold"] = expense.get("char_level_gold", 0) + gold_needed
            else:
                if resources.character_exp < exp_needed:
                    bottlenecks.append(f"角色经验不足 (需要{exp_needed}, 有{resources.character_exp})")
                if resources.gold < gold_needed:
                    bottlenecks.append(f"金币不足(角色升级) (需要{gold_needed}, 有{resources.gold})")
                break

        # 其次升级专武
        for lv, required_exp in sorted(WEAPON_LEVEL_EXP.items()):
            if lv <= progress.weapon_level:
                continue
            exp_needed = required_exp - progress.weapon_exp_used
            gold_needed = WEAPON_LEVEL_GOLD[lv] - progress.weapon_gold_used
            if resources.weapon_exp >= exp_needed and resources.gold >= gold_needed:
                resources.weapon_exp -= exp_needed
                resources.gold -= gold_needed
                progress.weapon_exp_used = required_exp
                progress.weapon_gold_used = WEAPON_LEVEL_GOLD[lv]
                progress.weapon_level = lv
                milestones.append(f"专武达到 Lv{lv}")
                expense["weap_level_exp"] = expense.get("weap_level_exp", 0) + exp_needed
                expense["weap_level_gold"] = expense.get("weap_level_gold", 0) + gold_needed
            else:
                if resources.weapon_exp < exp_needed:
                    bottlenecks.append(f"专武经验不足 (需要{exp_needed}, 有{resources.weapon_exp})")
                break

        # === 生成报告 ===
        report = DailyReport(
            day=day,
            income=income,
            expense=expense,
            totals={
                "gold": resources.gold,
                "diamond": resources.diamond,
                "character_exp": resources.character_exp,
                "weapon_exp": resources.weapon_exp,
                "character_level": progress.character_level,
                "weapon_level": progress.weapon_level,
                "mainline_stage": progress.mainline_stage,
            },
            milestones=milestones,
            bottlenecks=list(set(bottlenecks)),  # 去重
        )
        reports.append(report)

    return reports


def print_summary(reports: List[DailyReport], scenario_name: str):
    """打印模拟摘要"""
    print(f"\n{'='*60}")
    print(f"  场景: {scenario_name} | 模拟 {len(reports)} 天")
    print(f"{'='*60}")

    # 关键时间点
    checkpoints = [1, 3, 7, 14, 30, 60, 90]
    print(f"\n{'天数':<6}{'角色等级':<10}{'专武等级':<10}{'主线进度':<10}{'金币存量':<12}{'钻石存量':<10}")
    print("-" * 58)
    for cp in checkpoints:
        if cp <= len(reports):
            r = reports[cp - 1]
            t = r.totals
            print(f"Day {cp:<3}{t['character_level']:<10}{t['weapon_level']:<10}"
                  f"{t['mainline_stage']}/60{'':<5}{t['gold']:<12}{t['diamond']:<10}")

    # 里程碑时间线
    print(f"\n里程碑:")
    for r in reports:
        for m in r.milestones:
            print(f"  Day {r.day}: {m}")

    # 首次瓶颈
    print(f"\n首次出现瓶颈:")
    seen = set()
    for r in reports:
        for b in r.bottlenecks:
            key = b.split("(")[0].strip()
            if key not in seen:
                seen.add(key)
                print(f"  Day {r.day}: {b}")

    final = reports[-1].totals
    print(f"\n最终状态 (Day {len(reports)}):")
    print(f"  角色等级: Lv{final['character_level']}")
    print(f"  专武等级: Lv{final['weapon_level']}")
    print(f"  主线进度: {final['mainline_stage']}/60")
    print(f"  金币存量: {final['gold']:,}")
    print(f"  钻石存量: {final['diamond']:,}")
    print(f"  角色经验余量: {final['character_exp']:,}")
    print(f"  专武经验余量: {final['weapon_exp']:,}")


def main():
    global CHARACTER_LEVEL_EXP, CHARACTER_LEVEL_GOLD, WEAPON_LEVEL_EXP, WEAPON_LEVEL_GOLD, MAINLINE_REWARDS, MAINLINE_STAGE_COUNT

    parser = argparse.ArgumentParser(description="Vgame Player Progression Simulator")
    parser.add_argument("--days", type=int, default=30, help="模拟天数")
    parser.add_argument("--scenario", type=str, default="f2p_active",
                       help="场景名(逗号分隔多个进行对比): f2p_active, f2p_casual, light_spender, medium_spender")
    parser.add_argument("--compare", action="store_true", help="对比多个场景")
    parser.add_argument("--override", type=str, default="", help="参数覆盖，格式: key=value,key2=value2")
    parser.add_argument("--json", action="store_true", help="输出 JSON 格式")
    parser.add_argument("--from-excel", action="store_true", help="从 Vgame经济框架.xlsx 自动读取最新数据")
    parser.add_argument("--excel-path", type=str, default="", help="自定义经济框架 Excel 路径")
    args = parser.parse_args()

    # 从 Excel 加载最新数据
    if args.from_excel:
        if args.excel_path:
            global ECONOMY_XLSX_PATH
            ECONOMY_XLSX_PATH = args.excel_path
        result = load_from_excel()
        if result[0] is not None:
            CHARACTER_LEVEL_EXP, CHARACTER_LEVEL_GOLD, WEAPON_LEVEL_EXP, WEAPON_LEVEL_GOLD, MAINLINE_REWARDS, MAINLINE_STAGE_COUNT = result
        else:
            print("[WARNING] Excel 读取失败，继续使用内置默认值。")

    scenarios = [s.strip() for s in args.scenario.split(",")]

    for scenario_name in scenarios:
        params = get_scenario(scenario_name)

        # 应用参数覆盖
        if args.override:
            for override in args.override.split(","):
                key, value = override.split("=")
                if hasattr(params, key.strip()):
                    attr_type = type(getattr(params, key.strip()))
                    setattr(params, key.strip(), attr_type(value.strip()))

        reports = simulate(params, args.days)

        if args.json:
            output = {
                "scenario": scenario_name,
                "days": args.days,
                "reports": [
                    {
                        "day": r.day,
                        "totals": r.totals,
                        "milestones": r.milestones,
                        "bottlenecks": r.bottlenecks,
                    }
                    for r in reports
                ]
            }
            print(json.dumps(output, ensure_ascii=False, indent=2))
        else:
            print_summary(reports, scenario_name)


if __name__ == "__main__":
    main()
