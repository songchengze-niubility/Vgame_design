r"""
Vgame Config Schema Extractor
================================
从 Luban Excel 配置表提取 Schema，生成 config-schema.json。
零 token 消耗，纯 Python 结构化提取。

Usage:
    python extract_vgame_config_schema.py
    python extract_vgame_config_schema.py --datas-root "<config-datas-path>"

输入: ${VGAME_CONFIG_DATAS}/__tables__.xlsx + 所有注册的 .xlsx
输出: config-schema.json（配置表字段、类型、外键、公式列、C#引用）
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any, Optional

from vgame_paths import code_root, config_datas

# ========== 默认路径 ==========

DEFAULT_DATAS_ROOT = str(config_datas())
DEFAULT_CODE_ROOT = str(code_root())
SCRIPT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = str(SCRIPT_ROOT / "knowledge-graph" / "config-schema.json")
DEFAULT_FINGERPRINTS = str(SCRIPT_ROOT / "knowledge-graph" / "fingerprints.json")


def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)


def load_tables_registry(datas_root: str) -> list[dict]:
    """
    从 __tables__.xlsx 读取注册的逻辑表清单。
    返回: [{logical_name, input_files, ...}, ...]
    """
    openpyxl = ensure_openpyxl()
    tables_path = os.path.join(datas_root, "__tables__.xlsx")
    if not os.path.exists(tables_path):
        print(f"[ERROR] 未找到注册表: {tables_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(tables_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 找到 header 行（通常第1行是 ##var）
    tables = []
    header_row = None
    headers = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), start=1):
        first_cell = row[0].value
        if first_cell and str(first_cell).strip() == "##var":
            header_row = row_idx
            for col_idx, cell in enumerate(row):
                if cell.value:
                    headers[str(cell.value).strip()] = col_idx
            break
    
    if not header_row:
        # 尝试直接读第一行作为 header
        for col_idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        header_row = 1

    # 读数据行
    # 常见字段: full_name / name, input, ...
    name_col = headers.get("full_name", headers.get("name", headers.get("fullName", 0)))
    input_col = headers.get("input", headers.get("Input", None))
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        first_val = str(row[0]).strip()
        if first_val.startswith("##") or not first_val:
            continue
        
        logical_name = str(row[name_col]).strip() if name_col is not None and row[name_col] else ""
        input_files = str(row[input_col]).strip() if input_col is not None and len(row) > input_col and row[input_col] else ""
        
        if logical_name:
            tables.append({
                "logical_name": logical_name,
                "input_files": input_files,
            })
    
    wb.close()
    return tables


def find_excel_files(datas_root: str) -> list[Path]:
    """找到所有 .xlsx 文件（排除 ~$ 临时文件和 __tables__/__beans__/__enums__）"""
    result = []
    skip_prefixes = ("~$", "__")
    
    for xlsx_path in Path(datas_root).rglob("*.xlsx"):
        if xlsx_path.name.startswith(skip_prefixes):
            continue
        if xlsx_path.name.startswith("."):
            continue
        result.append(xlsx_path)
    return result


def extract_table_schema(xlsx_path: Path, datas_root: str) -> Optional[dict]:
    """
    从单个 .xlsx 文件提取 schema。
    读取前4行 header + 扫描公式。
    """
    openpyxl = ensure_openpyxl()
    
    try:
        # 第一遍: data_only=False 检测公式
        wb_formula = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=False)
        ws_formula = wb_formula.active
        
        # 读前4行确定结构
        rows_raw = []
        for i, row in enumerate(ws_formula.iter_rows(min_row=1, max_row=4, values_only=False), start=1):
            rows_raw.append(row)
            if i >= 4:
                break
        
        if len(rows_raw) < 2:
            wb_formula.close()
            return None
        
        # 判断哪一行是 ##var
        var_row_idx = None
        type_row_idx = None
        desc_row_idx = None
        
        for idx, row in enumerate(rows_raw):
            first_val = row[0].value if row[0] else None
            if first_val and str(first_val).strip() == "##var":
                var_row_idx = idx
            elif first_val and str(first_val).strip() == "##type":
                type_row_idx = idx
            elif first_val and str(first_val).strip() == "##":
                desc_row_idx = idx
        
        # 如果没有标准 Luban header，跳过
        if var_row_idx is None:
            wb_formula.close()
            return None
        
        # 提取字段名
        var_row = rows_raw[var_row_idx]
        field_names = []
        for cell in var_row:
            val = cell.value
            if val and str(val).strip() not in ("##var", "##type", "##group", "##"):
                field_names.append(str(val).strip())
            elif val and str(val).strip() in ("##var",):
                field_names.append("__tag__")
            else:
                field_names.append(None)
        
        # 提取字段类型
        field_types = []
        if type_row_idx is not None:
            type_row = rows_raw[type_row_idx]
            for cell in type_row:
                val = cell.value
                if val and str(val).strip() not in ("##type",):
                    field_types.append(str(val).strip())
                else:
                    field_types.append(None)
        
        # 提取中文描述
        field_descs = []
        if desc_row_idx is not None:
            desc_row = rows_raw[desc_row_idx]
            for cell in desc_row:
                val = cell.value
                if val and str(val).strip() not in ("##",):
                    field_descs.append(str(val).strip())
                else:
                    field_descs.append(None)
        
        # 扫描数据区前10行检测公式列
        formula_columns = set()
        data_start = (var_row_idx + 1) + 1  # 跳过所有 header 行后的第一行
        # 实际数据起始行 = max(var, type, group, desc) + 1，通常是第5行
        actual_data_start = max(var_row_idx, type_row_idx or 0, desc_row_idx or 0) + 2  # +2 因为 idx 从0开始
        
        for row_idx, row in enumerate(ws_formula.iter_rows(min_row=actual_data_start, max_row=actual_data_start + 10, values_only=False)):
            for col_idx, cell in enumerate(row):
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_columns.add(col_idx)
        
        wb_formula.close()
        
        # 构建字段列表
        fields = {}
        for col_idx in range(len(field_names)):
            fname = field_names[col_idx] if col_idx < len(field_names) else None
            if not fname or fname == "__tag__":
                continue
            
            ftype = field_types[col_idx] if col_idx < len(field_types) else None
            fdesc = field_descs[col_idx] if col_idx < len(field_descs) else None
            has_formula = col_idx in formula_columns
            
            fields[fname] = {
                "type": ftype,
                "description": fdesc,
                "column_index": col_idx,
                "has_formula": has_formula,
            }
        
        # 相对路径
        rel_path = str(xlsx_path.relative_to(datas_root)).replace("\\", "/")
        
        # Sheet 名
        wb_data = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        sheet_name = wb_data.active.title
        # 估算数据行数
        max_row = wb_data.active.max_row or 0
        data_rows = max(0, max_row - actual_data_start + 1)
        wb_data.close()
        
        return {
            "source_file": rel_path,
            "sheet_name": sheet_name,
            "fields": fields,
            "formula_columns": [field_names[i] for i in sorted(formula_columns) if i < len(field_names) and field_names[i] and field_names[i] != "__tag__"],
            "safe_write_columns": [fname for fname, fdata in fields.items() if not fdata["has_formula"] and fname != "__tag__"],
            "data_rows": data_rows,
        }
        
    except Exception as e:
        print(f"  [WARN] 跳过 {xlsx_path.name}: {e}")
        return None


def infer_foreign_keys(all_tables: dict[str, dict]) -> list[dict]:
    """
    推断外键关系。规则:
    - 字段名以 Id/ID 结尾 → 搜索是否存在对应逻辑表
    - 特殊映射: DropId → Drop, UnlockId → Unlock, LevelId → level
    """
    table_names_lower = {name.lower(): name for name in all_tables.keys()}
    edges = []
    
    # 已知特殊映射
    special_mappings = {
        "dropid": "Drop",
        "unlockid": "Unlock",
        "levelid": "level",
        "heroid": "Hero",
        "skillid": "Skill",
        "buffid": "Buff",
        "monsterid": "Monster",
        "itemid": "Items",
        "shopid": "Shop",
        "taskid": "Task",
        "activityid": "Activity",
        "chapterid": "ChapterLevel",
    }
    
    for table_name, table_data in all_tables.items():
        for field_name in table_data.get("fields", {}).keys():
            # 检查是否以 Id/ID 结尾
            if not (field_name.endswith("Id") or field_name.endswith("ID")):
                continue
            
            # 跳过主键（通常是第一个字段且叫 Id/ID）
            if field_name in ("Id", "ID", "id"):
                continue
            
            field_lower = field_name.lower()
            
            # 先查特殊映射
            target = special_mappings.get(field_lower)
            
            if not target:
                # 去掉 Id/ID 后缀，搜索表名
                if field_name.endswith("Id"):
                    candidate = field_name[:-2]
                elif field_name.endswith("ID"):
                    candidate = field_name[:-2]
                else:
                    continue
                
                # 精确匹配
                if candidate in all_tables:
                    target = candidate
                elif candidate.lower() in table_names_lower:
                    target = table_names_lower[candidate.lower()]
            
            if target and target in all_tables and target != table_name:
                edges.append({
                    "source": f"table:{table_name}",
                    "target": f"table:{target}",
                    "type": "depends_on",
                    "field": field_name,
                })
    
    return edges


def search_cs_references(table_name: str, code_root: str) -> list[str]:
    """
    在 C# 代码中搜索对配置表的引用。
    搜索模式: Tables.{TableName} 或 Get{TableName}
    """
    if not os.path.exists(code_root):
        return []
    
    refs = []
    patterns = [
        f"Tables.{table_name}",
        f"Get{table_name}",
        f'"{table_name}"',
    ]
    
    try:
        for cs_file in Path(code_root).rglob("*.cs"):
            if cs_file.name.startswith("~"):
                continue
            try:
                content = cs_file.read_text(encoding="utf-8", errors="ignore")
                for pattern in patterns:
                    if pattern in content:
                        rel = str(cs_file.relative_to(code_root)).replace("\\", "/")
                        refs.append(f"{rel}")
                        break
            except:
                continue
    except:
        pass
    
    return refs[:10]  # 最多返回10个引用


def compute_fingerprints(datas_root: str, xlsx_files: list[Path]) -> dict:
    """计算文件 hash 指纹"""
    fps = {}
    for f in xlsx_files:
        try:
            h = hashlib.sha256(f.read_bytes()).hexdigest()
            rel = str(f.relative_to(datas_root)).replace("\\", "/")
            fps[rel] = {"hash": h, "size": f.stat().st_size}
        except:
            pass
    return fps


def main():
    parser = argparse.ArgumentParser(description="Vgame Config Schema Extractor")
    parser.add_argument("--datas-root", default=DEFAULT_DATAS_ROOT, help="Datas 目录路径")
    parser.add_argument("--code-root", default=DEFAULT_CODE_ROOT, help="C# 代码根目录")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="输出 JSON 路径")
    parser.add_argument("--fingerprints", default=DEFAULT_FINGERPRINTS, help="指纹文件路径")
    parser.add_argument("--skip-cs-refs", action="store_true", help="跳过 C# 引用搜索（加速）")
    args = parser.parse_args()

    start_time = time.time()
    print(f"[INFO] Vgame Config Schema Extractor")
    print(f"[INFO] Datas 目录: {args.datas_root}")
    
    # 1. 找到所有 Excel 文件
    xlsx_files = find_excel_files(args.datas_root)
    print(f"[INFO] 发现 {len(xlsx_files)} 个 .xlsx 文件")
    
    # 2. 读取注册表（可选，用于映射逻辑表名）
    registry = []
    try:
        registry = load_tables_registry(args.datas_root)
        print(f"[INFO] __tables__.xlsx 注册了 {len(registry)} 个逻辑表")
    except Exception as e:
        print(f"[WARN] 读取 __tables__.xlsx 失败: {e}，将使用文件名作为表名")
    
    # 逻辑表名到源文件的映射
    logical_to_files = {}
    for reg in registry:
        logical_to_files[reg["logical_name"]] = reg.get("input_files", "")
    
    # 3. 逐文件提取 Schema
    all_tables = {}
    skipped = 0
    
    for i, xlsx_path in enumerate(xlsx_files):
        if (i + 1) % 50 == 0:
            print(f"  提取中... {i+1}/{len(xlsx_files)}")
        
        schema = extract_table_schema(xlsx_path, args.datas_root)
        if schema:
            # 用文件名（不含后缀）作为表名，除非有注册名
            table_name = xlsx_path.stem
            all_tables[table_name] = schema
        else:
            skipped += 1
    
    print(f"[INFO] 成功提取 {len(all_tables)} 张表，跳过 {skipped} 个文件")
    
    # 4. 推断外键关系
    edges = infer_foreign_keys(all_tables)
    print(f"[INFO] 推断出 {len(edges)} 条外键关系")
    
    # 5. 搜索 C# 引用（可选）
    if not args.skip_cs_refs and os.path.exists(args.code_root):
        print(f"[INFO] 搜索 C# 代码引用...")
        for table_name in all_tables:
            refs = search_cs_references(table_name, args.code_root)
            if refs:
                all_tables[table_name]["cs_refs"] = refs
        cs_count = sum(1 for t in all_tables.values() if t.get("cs_refs"))
        print(f"[INFO] {cs_count} 张表找到 C# 引用")
    else:
        print(f"[INFO] 跳过 C# 引用搜索")
    
    # 6. 构建输出
    nodes = []
    for table_name, schema in all_tables.items():
        nodes.append({
            "id": f"table:{table_name}",
            "type": "table",
            "layer": "config",
            "name": table_name,
            "metadata": schema,
        })
    
    output = {
        "version": "1.0",
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "datas_root": "${VGAME_CONFIG_DATAS}",
        "statistics": {
            "total_tables": len(all_tables),
            "total_fields": sum(len(t.get("fields", {})) for t in all_tables.values()),
            "total_edges": len(edges),
            "tables_with_formulas": sum(1 for t in all_tables.values() if t.get("formula_columns")),
            "tables_with_cs_refs": sum(1 for t in all_tables.values() if t.get("cs_refs")),
        },
        "nodes": nodes,
        "edges": edges,
    }
    
    # 7. 写入
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    output_size = os.path.getsize(args.output)
    print(f"[INFO] 输出: {args.output} ({output_size / 1024:.1f} KB)")
    
    # 8. 保存指纹
    fps = compute_fingerprints(args.datas_root, xlsx_files)
    os.makedirs(os.path.dirname(args.fingerprints), exist_ok=True)
    with open(args.fingerprints, "w", encoding="utf-8") as f:
        json.dump(fps, f, ensure_ascii=False, indent=2)
    print(f"[INFO] 指纹: {args.fingerprints} ({len(fps)} 文件)")
    
    # 9. 统计
    elapsed = time.time() - start_time
    stats = output["statistics"]
    print(f"\n{'='*50}")
    print(f"  提取完成 ({elapsed:.1f}s)")
    print(f"{'='*50}")
    print(f"  表数量:       {stats['total_tables']}")
    print(f"  字段总数:     {stats['total_fields']}")
    print(f"  外键关系:     {stats['total_edges']}")
    print(f"  含公式的表:   {stats['tables_with_formulas']}")
    print(f"  有C#引用的表: {stats['tables_with_cs_refs']}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
